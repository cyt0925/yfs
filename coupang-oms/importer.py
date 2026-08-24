"""整合表解析與差異比對。

比對規則（對應 OP 的實際情境）：
  - 資料庫查無 (PO, SKU)                    -> 新增
  - 存在且所有酷澎欄位相同                    -> 靜默略過（但要計數）
  - 存在但數量/交期/倉別等被改                -> 更新 + 欄位級歷程 + 標記待確認
  - 資料庫有、本次檔案沒有                    -> 完全不動（上傳的是片段，消失 != 取消）

「拆單」在酷澎會產生新的 PO 單號，因此走的是「新增」路徑，
不需要任何群組推測邏輯；只用 parent_po 記錄血緣。
"""

import openpyxl

from normalize import (
    norm_date, norm_decimal, norm_header, norm_int, norm_key, norm_text,
    norm_warehouse,
)

# 整合表欄位 -> 內部欄位。每個內部欄位可接受多個表頭別名，
# 因為酷澎/內部報表偶爾會微調標題文字。
COLUMN_ALIASES = {
    "po_number":    ["PO單號", "PO單號(採購單號)", "採購單號", "POID", "PO ID"],
    "sku_id":       ["SKUID", "SKU ID", "SKUNo.", "SKU No.", "SKU編號"],
    "order_type":   ["訂單類型", "PO類型"],
    "parent_po":    ["原訂單拆出", "原訂單", "母單號", "原PO單號"],
    "line":         ["線別"],
    "brand":        ["品牌"],
    "product_name": ["品名"],
    "barcode":      ["條碼(國條)", "條碼", "國條"],
    "yf_sku":       ["永豐料號", "料號"],
    "warehouse":    ["到貨倉別", "倉別", "FC"],
    "address":      ["地址"],
    "delivery_date": ["交付日期", "交貨日", "交期"],
    "qty_coupang":  ["下單數量(酷澎單位)", "下單數量", "PO數量"],
    "qty_file_ship": ["出貨數量"],
    "unit":         ["單位"],
    "box_size":     ["箱入數"],
    "unit_price":   ["酷澎下單單價(含稅)", "下單單價"],
    "expiry_note":  ["允收效期"],
    "seq_no":       ["NO", "No.", "項次"],
    "remarks_file": ["備註"],
}

# 酷澎擁有的欄位：只有匯入能改。任一變動都要記歷程並亮燈。
COUPANG_FIELDS = {
    "order_type":    "訂單類型",
    "parent_po":     "原訂單",
    "line":          "線別",
    "brand":         "品牌",
    "product_name":  "品名",
    "barcode":       "條碼",
    "yf_sku":        "永豐料號",
    "warehouse":     "倉別",
    "address":       "地址",
    "delivery_date": "交期",
    "qty_coupang":   "下單數量",
    "unit":          "單位",
    "box_size":      "箱入數",
    "unit_price":    "下單單價",
    "expiry_note":   "允收效期",
}

# 這幾個欄位一旦被改，就是會直接影響倉庫出貨的關鍵異動
CRITICAL_FIELDS = ("qty_coupang", "delivery_date", "warehouse")

REQUIRED = ("po_number", "sku_id")


class ImportError_(Exception):
    """匯入失敗。訊息會直接顯示給 OP，所以要寫人看得懂的話。"""


def _build_header_map(header_row):
    """把實際表頭對應到內部欄位名。"""
    normalized = {}
    for idx, cell in enumerate(header_row):
        key = norm_header(cell)
        if key and key not in normalized:
            normalized[key] = idx

    mapping = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            key = norm_header(alias)
            if key in normalized:
                mapping[field] = normalized[key]
                break
    return mapping


def _find_header_row(rows, max_scan=15):
    """表頭不一定在第 1 列，掃前幾列找出同時含 PO 與 SKU 的那一列。"""
    best = None
    for i, row in enumerate(rows[:max_scan]):
        mapping = _build_header_map(row)
        if all(f in mapping for f in REQUIRED):
            return i, mapping
        if best is None and len(mapping) >= 3:
            best = (i, mapping)
    if best:
        return best
    return None, {}


def parse_workbook(file_stream, filename=""):
    """讀取整合表，回傳 (rows, warnings)。rows 已完成正規化。"""
    try:
        wb = openpyxl.load_workbook(file_stream, data_only=True)
    except Exception as exc:  # noqa: BLE001 - 要把原因原樣告訴 OP
        raise ImportError_(f"無法開啟 Excel 檔案：{exc}") from exc

    sheet = None
    for name in wb.sheetnames:
        if "整合" in name:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb[wb.sheetnames[0]]

    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise ImportError_("這個檔案是空的，沒有任何資料列。")

    header_idx, mapping = _find_header_row(raw_rows)
    if header_idx is None or not all(f in mapping for f in REQUIRED):
        found = [norm_header(c) for c in raw_rows[0] if norm_header(c)][:12]
        raise ImportError_(
            "找不到「PO單號」與「SKU ID」欄位，這可能不是整合表。"
            f"（第一列讀到的欄位：{', '.join(found) or '空白'}）"
        )

    def cell(row, field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    rows, warnings = [], []
    for offset, raw in enumerate(raw_rows[header_idx + 1:], start=header_idx + 2):
        po = norm_key(cell(raw, "po_number"))
        sku = norm_key(cell(raw, "sku_id"))
        if not po and not sku:
            continue  # 空白列直接跳過，不算錯
        if not po or not sku:
            warnings.append(f"第 {offset} 列缺少 PO 單號或 SKU ID，已略過。")
            continue

        delivery_raw = cell(raw, "delivery_date")
        delivery = norm_date(delivery_raw)
        if delivery_raw not in (None, "") and not delivery:
            warnings.append(
                f"第 {offset} 列（{po} / {sku}）交期「{norm_text(delivery_raw)}」無法解析，已存為空白。"
            )

        qty = norm_int(cell(raw, "qty_coupang"))
        if qty is None:
            warnings.append(f"第 {offset} 列（{po} / {sku}）下單數量無法解析，已存為空白。")

        rows.append({
            "row_no": offset,
            "po_number": po,
            "sku_id": sku,
            "order_type": norm_text(cell(raw, "order_type")),
            "parent_po": norm_key(cell(raw, "parent_po")),
            "line": norm_text(cell(raw, "line")),
            "brand": norm_text(cell(raw, "brand")),
            "product_name": norm_text(cell(raw, "product_name")),
            "barcode": norm_key(cell(raw, "barcode")),
            "yf_sku": norm_key(cell(raw, "yf_sku")),
            "warehouse": norm_warehouse(cell(raw, "warehouse")),
            "address": norm_text(cell(raw, "address")),
            "delivery_date": delivery,
            "qty_coupang": qty,
            "qty_file_ship": norm_int(cell(raw, "qty_file_ship")),
            "unit": norm_text(cell(raw, "unit")),
            "box_size": norm_int(cell(raw, "box_size")),
            "unit_price": norm_decimal(cell(raw, "unit_price")),
            "expiry_note": norm_text(cell(raw, "expiry_note")),
            "seq_no": norm_int(cell(raw, "seq_no")),
            "remarks_file": norm_text(cell(raw, "remarks_file")),
            "source_file": filename,
        })

    if not rows:
        raise ImportError_("檔案裡沒有任何有效的訂單資料列。")

    # 同一個檔案內若出現重複的 (PO, SKU)，後面那筆會蓋掉前面那筆，
    # 這通常代表來源檔本身有問題，要讓 OP 知道。
    seen = {}
    dedup = []
    for row in rows:
        key = (row["po_number"], row["sku_id"])
        if key in seen:
            warnings.append(
                f"檔案內第 {row['row_no']} 列與第 {seen[key]} 列的 "
                f"{row['po_number']} / {row['sku_id']} 重複，採用後者。"
            )
            dedup = [r for r in dedup if (r["po_number"], r["sku_id"]) != key]
        seen[key] = row["row_no"]
        dedup.append(row)

    return dedup, warnings


def _display(field, value):
    if value is None or value == "":
        return "（空白）"
    return str(value)


def desired_ship(row):
    """整合表如果自己就帶了「出貨數量」欄，那才是 OP 真正要出的量，
    比死板地拿「下單數量」複製過去準——酷澎談好折讓、少出貨這些，
    整合表裡的出貨數量欄位會先反映出來。檔案沒帶這欄（None）才退回
    用下單數量頂著。"""
    file_ship = row.get("qty_file_ship")
    return file_ship if file_ship is not None else row.get("qty_coupang")


def diff_rows(conn, rows):
    """比對檔案與資料庫，回傳分類結果。此函式不寫任何資料。"""
    keys = [(r["po_number"], r["sku_id"]) for r in rows]
    existing = {}
    cur = conn.cursor()
    # 分批查，避免 SQLite 變數上限
    for i in range(0, len(keys), 300):
        chunk = keys[i:i + 300]
        placeholders = ",".join(["(?,?)"] * len(chunk))
        flat = [v for pair in chunk for v in pair]
        cur.execute(
            f"SELECT * FROM order_rows WHERE (po_number, sku_id) IN ({placeholders})",
            flat,
        )
        for db_row in cur.fetchall():
            existing[(db_row["po_number"], db_row["sku_id"])] = dict(db_row)

    new_rows, updated, identical = [], [], []

    for row in rows:
        key = (row["po_number"], row["sku_id"])
        old = existing.get(key)
        if old is None:
            new_rows.append(row)
            continue

        changes = []
        for field, label in COUPANG_FIELDS.items():
            old_val = old.get(field)
            new_val = row.get(field)
            # 檔案沒帶到的欄位（None / 空字串）不覆蓋既有值，
            # 避免來源報表少一欄就把資料庫洗白。
            if new_val is None or new_val == "":
                continue
            if _same(old_val, new_val):
                continue
            changes.append({
                "field": field,
                "label": label,
                "old": _display(field, old_val),
                "new": _display(field, new_val),
                "critical": field in CRITICAL_FIELDS,
            })

        # 出貨數量不是酷澎欄位（COUPANG_FIELDS 沒有它，因為資料庫裡的
        # qty_ship 名稱跟整合表的「出貨數量」欄不是同一件事、不能直接
        # 拿同名比對），所以獨立算一次：現在整合表算出來「應該要出多少」
        # 跟資料庫現在存的出貨數量，兩者是否不一樣。
        ship = desired_ship(row)
        ship_changed = ship is not None and not _same(old.get("qty_ship"), ship)

        if not changes and not ship_changed:
            identical.append(row)
        else:
            updated.append({
                "row": row,
                "order_id": old["id"],
                "version": old["version"],
                "is_pulled": bool(old["is_pulled"]),
                "qty_ship": old["qty_ship"],
                "qty_ship_overridden": bool(old["qty_ship_overridden"]),
                "changes": changes,
                "desired_ship": ship,
                "ship_changed": ship_changed,
                "after_pull": bool(old["is_pulled"]) and (
                    any(c["critical"] for c in changes) or ship_changed),
            })

    return {"new": new_rows, "updated": updated, "identical": identical}


def _same(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if isinstance(a, (int, float)) or isinstance(b, (int, float)):
        try:
            return abs(float(a) - float(b)) < 1e-9
        except (TypeError, ValueError):
            return str(a) == str(b)
    return str(a).strip() == str(b).strip()
