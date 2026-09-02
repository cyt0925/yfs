"""採購表轉換：跟酷澎訂單管理系統完全獨立的新功能。

轉換成各線別自己要用的「產品採購表」格式。三個線別各自沿用公司既有的
範本檔（purchase_templates/ 底下那三個 .xls，連查表工作表、儲存格樣式
都原封不動保留），只覆寫資料列——這樣匯出的檔案結構才會跟公司系統
原本要吃的格式一模一樣，不會因為改用別的函式庫重新產生而跑掉。

但三個線別的「匯入來源」長得完全不一樣，不是同一份檔案：
- P&G／紙潔：讀酷澎系統匯出的「訂單匯入」Excel（一列一個 SKU 的表格，
  見 parse_import_file）。
- 瑪氏：讀 MARS TAIWAN 開的「訂貨通知單」（ORDER FORM，一份很大的表，
  大部分列是「有賣過但這次沒訂」的參考列，只有填了採購數量的才是真的
  要訂的，見 parse_mars_order_form）。這份表的結構、來源都跟另外兩條線
  的匯入檔沒有任何關係，第一版把瑪氏也硬套進 P&G／紙潔那套解析邏輯是
  誤把同事給錯的範例當成正確格式，後來拿到真正的訂貨通知單才發現整個
  格式都是錯的。而且訂貨通知單一份就是一張 PO，所以瑪氏的上傳介面
  允許一次選多個檔案，各自轉成各自的一組。

不寫進資料庫、不碰任何一張酷澎訂單的表：上傳的檔案只在這次請求內
處理、解析結果全部回傳給瀏覽器暫存，使用者確認/編輯完再送回來產生
下載檔，處理完就丟，跟訂單管理系統的資料完全隔離。
"""
import io
import json
import os
import re
import zipfile

import openpyxl
import xlrd
from flask import Blueprint, jsonify, render_template, request, send_file
from xlutils.copy import copy as xl_copy

import db

purchase_bp = Blueprint("purchase", __name__)

TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), "purchase_templates")

# wipe_rows：範本檔主表原本的資料區域大概到第幾列（含預先編號的空白列），
# 產生新檔案時全部清成空白再重寫，兩份範例檔都留了不少預先編號、單位
# 已經填好「箱」的空白列在後面，這裡抓大一點，寧可多清幾列不要漏。
LINES = {
    "pg": {
        "label": "P&G",
        "template": os.path.join(TEMPLATE_DIR, "pg_template.xls"),
        "group_by": "po_number",
        "wipe_rows": 250,
        "default_remark_style": "buyout",
    },
    "paper": {
        "label": "紙潔",
        "template": os.path.join(TEMPLATE_DIR, "paper_template.xls"),
        "group_by": "po_number",
        "wipe_rows": 45,
        "default_remark_style": "blank",
    },
    "mars": {
        "label": "瑪氏",
        "template": os.path.join(TEMPLATE_DIR, "mars_template.xls"),
        "wipe_rows": 60,
        "default_remark_style": "mars",
        "multi_file": True,   # 一份訂貨通知單就是一張 PO，允許一次選多個檔案
    },
}

# 欄位名稱比對用——酷澎匯出的「訂單匯入」檔在 P&G／紙潔／瑪氏三條線
# 欄位順序不一樣（瑪氏把出貨備註搬到最前面），一律用標題文字找欄位，
# 不能寫死欄位字母位置。
FIELD_HEADERS = {
    "po_number": ["訂單編號"],
    "material_no": ["料號"],
    "qty": ["數量"],          # 前綴比對，因為標題後面偶爾會黏著總和數字
    "ship_note": ["出貨備註"],
    "address": ["收件地址"],
}

WAREHOUSE_FILE = "purchase_warehouses.json"
DEFAULT_WAREHOUSES = {
    "桃園市楊梅區環東路200號C區4樓": "TXRC8",
    "桃園市大園區建國路102號4樓 B棟(倉儲棟)": "TAO1",
    "桃園市觀音區玉林路一段523號": "TAO4",
}


def _load_warehouses():
    path = os.path.join(db.DATA_DIR, WAREHOUSE_FILE)
    try:
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
            if isinstance(data, dict) and data:
                return data
    except (OSError, ValueError):
        pass
    return dict(DEFAULT_WAREHOUSES)


def _save_warehouses(mapping):
    os.makedirs(db.DATA_DIR, exist_ok=True)
    path = os.path.join(db.DATA_DIR, WAREHOUSE_FILE)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(mapping, fh, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _find_column(headers, candidates):
    for i, h in enumerate(headers):
        if h is None:
            continue
        text = str(h).strip()
        for cand in candidates:
            if text == cand or text.startswith(cand):
                return i
    return None


def _guess_date_mmdd(ship_note):
    """從出貨備註開頭抓 MMDD（例如「0905從PG出貨至TXRC8」→ "0905"）。"""
    if not ship_note:
        return ""
    m = re.match(r"^(\d{2})(\d{2})", ship_note.strip())
    if not m:
        return ""
    month, day = int(m.group(1)), int(m.group(2))
    if 1 <= month <= 12 and 1 <= day <= 31:
        return m.group(0)
    return ""


def _guess_date_from_filename(filename):
    """出貨備註猜不到日期時（瑪氏就是這樣，出貨備註是內部採購單號，
    不含日期）退而看上傳檔名——三個線別的匯入檔名都會帶著「MMDD 到貨」
    這組數字，例如「酷澎訂單匯入_0904到貨_TAO1_TAO4.xlsx」。

    但「到貨」不是唯一的講法，實測遇過「交貨」（酷澎訂單匯入_0904交貨-
    TAO1、TAO4.xlsx），所以「到／交／出」都認，不要只認死一種寫法，
    不然換個人上傳、換個講法就又猜不到。"""
    if not filename:
        return ""
    m = re.search(r"(\d{2})(\d{2})[到交出]貨", filename)
    if not m:
        return ""
    month, day = int(m.group(1)), int(m.group(2))
    if 1 <= month <= 12 and 1 <= day <= 31:
        return m.group(1) + m.group(2)
    return ""


def _guess_po_from_filename(filename):
    """瑪氏的訂貨通知單裡「永豐PO單號」那個欄位常常是空的（見樣本），
    訂單編號改從上傳的檔名抓——酷澎訂單編號固定是 13 開頭的 14 碼數字，
    例如「13000000467952」。跟猜日期同一個道理：檔名裡有就用，抓不到
    就留空給人填，不硬湊。"""
    if not filename:
        return ""
    # 不能用 \b 當邊界——檔名裡數字前後常常接底線或中文字，這兩種在
    # regex 裡都算「單字字元」，跟數字之間沒有邊界，\b 會直接抓不到。
    # 改用「前後不是數字」，這樣不管接的是底線、中文、副檔名都擋得住，
    # 也不會不小心咬到更長數字串裡的一段。
    m = re.search(r"(?<!\d)(13\d{12})(?!\d)", filename)
    return m.group(1) if m else ""


def mmdd_to_md(mmdd):
    """0905 -> 9/5（備註文字慣用不補零的月/日）。"""
    if not mmdd or len(mmdd) != 4 or not mmdd.isdigit():
        return ""
    return f"{int(mmdd[:2])}/{int(mmdd[2:])}"


def _guess_warehouse_from_note(ship_note):
    """PG 的出貨備註格式是「MMDD從PG出貨至TXRC8」，倉別直接寫在裡面；
    其他線別的出貨備註格式不含倉別，抓不到就回傳空字串，交給地址對照表。"""
    if not ship_note:
        return ""
    m = re.search(r"至([A-Z0-9]+)\s*$", ship_note.strip())
    return m.group(1) if m else ""


class PurchaseImportError(Exception):
    pass


_MARS_HEADER_SCAN_ROWS = 15    # 表頭資訊區塊大概落在這幾列裡
_MARS_HEADER_SCAN_COLS = 12
CATEGORY_LIST_SHEET = "清單"

# 品類全名（「清單」分頁裡的選項）→ 檔名裡實際要用的品類代碼。這是公司
# 既有檔名慣例定的固定寫法，不是從全名機械推導出來的（例如 Chocolate
# 巧克力縮寫成「Cho巧」，不是取前三碼），所以用明確對照表，不要用猜的。
CATEGORY_CODE_MAP = {
    "Petcare寵物": "PET寵物",
    "Gum糖果": "GUM糖",
    "Chocolate巧克力": "Cho巧",
}


def _find_label_value(ws, label, max_row=_MARS_HEADER_SCAN_ROWS,
                       max_col=_MARS_HEADER_SCAN_COLS):
    """訂貨通知單的表頭是「標籤格＋同一列右邊某格是值」這種鬆散排版
    （例如「配送日: 」在 F3、值在 H3，中間 G3 是空的），不是固定的表格，
    找值要用「先找標籤文字、再往右找第一個有東西的格」，不能寫死座標。"""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            text = str(v).strip().rstrip(":：").strip()
            if text == label:
                for c2 in range(c + 1, max_col + 1):
                    v2 = ws.cell(row=r, column=c2).value
                    if v2 is not None and str(v2).strip() != "":
                        return v2
                return None
    return None


def _find_mars_data_sheet(wb):
    """訂貨通知單通常有兩個分頁（一個是表本身，一個是給下拉選單用的
    品類清單），用哪個當主表不能寫死分頁名稱（連現有樣本的分頁名稱都
    帶著一個容易漏看的尾隨空白），改成找「哪個分頁裡有一格剛好等於
    永豐料號」，那就是主表。"""
    for ws in wb.worksheets:
        for row in ws.iter_rows(max_row=30, values_only=True):
            if "永豐料號" in row:
                return ws
    return None


def _guess_mars_category_code(wb, data_ws):
    """檔名裡「GUM糖」「PET寵物」「Cho巧」這段是品類代碼，來源是主表裡
    某一格直接寫著品類全名（例如「Gum糖果」）。優先讀「清單」分頁列出的
    合法品類全名去比對主表，抓到全名後查 CATEGORY_CODE_MAP 換成公司
    檔名慣例的代碼；查不到對照（例如清單以後新增了品類但對照表還沒更新）
    就退回取英文開頭三碼當備援猜測，抓不到就回傳空字串，讓使用者自己填，
    不要猜錯。"""
    categories = []
    if CATEGORY_LIST_SHEET in wb.sheetnames:
        for row in wb[CATEGORY_LIST_SHEET].iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str) and v.strip():
                    categories.append(v.strip())

    def to_code(full_text):
        if full_text in CATEGORY_CODE_MAP:
            return CATEGORY_CODE_MAP[full_text]
        m = re.match(r"^([A-Za-z]+)", full_text)
        return m.group(1)[:3].upper() if m else ""

    if categories:
        cat_set = set(categories)
        for row in data_ws.iter_rows(max_row=15, values_only=True):
            for v in row:
                if isinstance(v, str) and v.strip() in cat_set:
                    return to_code(v.strip())

    for row in data_ws.iter_rows(max_row=15, values_only=True):
        for v in row:
            if isinstance(v, str) and re.match(r"^[A-Za-z]+[一-鿿]", v.strip()):
                return to_code(v.strip())
    return ""


def parse_mars_order_form(file_bytes, filename=""):
    """讀 MARS TAIWAN 的「訂貨通知單」（ORDER FORM），回傳單一組資料
    （一份訂貨通知單就是一張 PO，不像 P&G／紙潔那樣要分組）。

    表格結構分兩塊：
    - 表頭資訊區（前面十幾列）：配送日、入倉倉別、送貨地址、永豐PO單號、
      品類，用標籤文字去找值（見 _find_label_value）。
    - 資料區：從「永豐料號」那一列表頭開始，一路列到底，但大部分列是
      「有賣過、這次沒訂」的參考列，只有「採購數量」欄有填數字的才是
      真的要訂的品項，其他要濾掉，不能整批照抓。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise PurchaseImportError(f"檔案讀取失敗，請確認是有效的 Excel 檔：{exc}") from exc

    ws = _find_mars_data_sheet(wb)
    if ws is None:
        raise PurchaseImportError(
            "找不到「永豐料號」這個欄位，請確認上傳的是瑪氏的訂貨通知單（ORDER FORM）。")

    header_row_idx = None
    material_col = qty_col = None
    for row in ws.iter_rows(max_row=30):
        for cell in row:
            if cell.value == "永豐料號":
                header_row_idx = cell.row
                material_col = cell.column
                break
        if header_row_idx:
            break
    for cell in ws[header_row_idx]:
        if cell.value and str(cell.value).strip().startswith("採購數量"):
            qty_col = cell.column
            break
    if qty_col is None:
        raise PurchaseImportError("找不到「採購數量」欄位，請確認上傳的是瑪氏的訂貨通知單。")

    rows = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        material = ws.cell(row=r, column=material_col).value
        if material is None or str(material).strip() == "":
            continue
        raw_qty = ws.cell(row=r, column=qty_col).value
        try:
            qty = int(raw_qty)
        except (TypeError, ValueError):
            continue
        if qty == 0:
            continue
        rows.append({"material_no": str(material).strip(), "qty": qty})

    if not rows:
        raise PurchaseImportError(
            "這份訂貨通知單裡找不到任何有填「採購數量」的品項，請確認上傳的檔案有勾選數量。")

    delivery_date = _find_label_value(ws, "配送日")
    date_guess = ""
    if hasattr(delivery_date, "year"):
        date_guess = f"{delivery_date.month:02d}{delivery_date.day:02d}"
    if not date_guess:
        date_guess = _guess_date_from_filename(filename)

    warehouse_raw = _find_label_value(ws, "入倉倉別") or ""
    m = re.search(r"([A-Za-z0-9]+)\s*$", str(warehouse_raw).strip())
    warehouse_guess = m.group(1).upper() if m else ""

    address = str(_find_label_value(ws, "送貨地址") or "").strip()

    po_raw = str(_find_label_value(ws, "永豐PO單號") or "").strip()
    po_number = po_raw if re.fullmatch(r"\d{10,}", po_raw) else _guess_po_from_filename(filename)

    category_code = _guess_mars_category_code(wb, ws)

    filename_stem = os.path.splitext(filename)[0] if filename else ""
    key = po_number or (filename_stem or f"檔案{id(file_bytes) % 10000}")

    return {
        "key": key,
        "po_numbers": [po_number] if po_number else [],
        "address": address,
        "address_mismatch": False,
        "warehouse_guess": warehouse_guess,
        "date_guess": date_guess,
        "category": category_code,
        "item_count": len(rows),
        "qty_total": sum(r["qty"] for r in rows),
        "rows": rows,
        # 匯出檔名直接沿用匯入檔名（同事反映匯出後檔名對不起來，不容易
        # 找到自己剛剛匯入的是哪一份）；副檔名固定改 .xls，因為匯出的
        # 一定是舊版 BIFF 格式，副檔名跟著原檔名走的話（例如原檔是
        # .xlsx）打開會被 Excel 警告格式不符。
        "filename_stem": filename_stem,
    }


def parse_import_file(line_key, file_bytes, filename=""):
    """讀酷澎的「訂單匯入」Excel，依線別規則分組，回傳給前端預覽用的結構。

    filename 是上傳時的原始檔名，只用來在出貨備註猜不到日期時當備援
    （見 _guess_date_from_filename）——不影響解析內容，純粹輔助猜日期。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as exc:
        raise PurchaseImportError(f"檔案讀取失敗，請確認是有效的 Excel 檔：{exc}") from exc

    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if header_row is None:
        raise PurchaseImportError("檔案是空的，沒有表頭列。")
    headers = [c.value for c in header_row]

    cols = {}
    missing = []
    for field, candidates in FIELD_HEADERS.items():
        idx = _find_column(headers, candidates)
        if idx is None:
            missing.append(candidates[0])
        cols[field] = idx
    if missing:
        raise PurchaseImportError(
            f"檔案裡找不到欄位：{'、'.join(missing)}，請確認上傳的是酷澎「訂單匯入」檔。")

    warehouses = _load_warehouses()
    line_cfg = LINES[line_key]

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        po = r[cols["po_number"]]
        material = r[cols["material_no"]]
        if po is None or material is None or str(po).strip() == "":
            continue
        po = str(po).strip()
        material = str(material).strip()
        raw_qty = r[cols["qty"]]
        try:
            qty = int(raw_qty)
        except (TypeError, ValueError):
            qty = 0
        ship_note = str(r[cols["ship_note"]] or "").strip() if cols["ship_note"] is not None else ""
        address = str(r[cols["address"]] or "").strip() if cols["address"] is not None else ""
        rows.append({
            "po_number": po, "material_no": material, "qty": qty,
            "ship_note": ship_note, "address": address,
        })

    if not rows:
        raise PurchaseImportError("這個檔案解析不到任何資料列，請確認上傳的是酷澎「訂單匯入」檔。")

    group_field = line_cfg["group_by"]
    groups_order = []
    groups = {}
    for row in rows:
        key = row[group_field] or row["po_number"]
        if key not in groups:
            groups[key] = []
            groups_order.append(key)
        groups[key].append(row)

    result = []
    for key in groups_order:
        grows = groups[key]
        po_numbers = sorted({r["po_number"] for r in grows})
        addresses = {r["address"] for r in grows if r["address"]}
        address = next(iter(addresses)) if len(addresses) == 1 else ""

        warehouse_guess = warehouses.get(address, "") if address else ""
        if not warehouse_guess:
            for r in grows:
                warehouse_guess = _guess_warehouse_from_note(r["ship_note"])
                if warehouse_guess:
                    break

        date_guess = ""
        for r in grows:
            date_guess = _guess_date_mmdd(r["ship_note"])
            if date_guess:
                break
        if not date_guess:
            # 瑪氏的出貨備註就是內部採購單號，不含日期，退而看上傳檔名——
            # 三個線別的匯入檔名都遵守同一個「MMDD到貨」慣例。
            date_guess = _guess_date_from_filename(filename)

        result.append({
            "key": str(key),
            "po_numbers": po_numbers,
            "address": address,
            "address_mismatch": len(addresses) > 1,
            "warehouse_guess": warehouse_guess,
            "date_guess": date_guess,
            "item_count": len(grows),
            "qty_total": sum(r["qty"] for r in grows),
            "rows": [{"material_no": r["material_no"], "qty": r["qty"]} for r in grows],
        })
    return result


def build_remark(line_key, po_numbers, date_mmdd, warehouse):
    """三個線別各自的備註樣板。日期／倉別任一沒填就先留空，讓使用者自己補，
    不要硬湊出一句看起來對、其實缺資料的備註。"""
    style = LINES[line_key]["default_remark_style"]
    md = mmdd_to_md(date_mmdd)
    if style == "blank":
        return ""
    if style == "buyout":
        po = po_numbers[0] if po_numbers else ""
        if not (po and warehouse and md):
            return ""
        return f"{po}({warehouse})_酷澎{md}買斷"
    if style == "mars":
        if not (md and warehouse):
            return ""
        return f"MARS入倉{md}瑪氏送酷澎-{warehouse}倉"
    return ""


def build_filename(line_key, po_numbers, date_mmdd, warehouse, category=""):
    """檔名裡的 PO 單號是用來區分「一張 PO 一個檔」時是哪一張，所以合併
    匯出時呼叫端會傳空的 po_numbers 進來，這裡就不放單號——這正是紙潔
    （本來就是多張 PO 合併）的真實檔名慣例：酷澎_產品採購表上傳_0902到貨。

    category 是瑪氏檔名裡「GUM糖」「PET寵物」「Cho巧」那段（品類代碼，
    來自訂貨通知單裡的品類全名對照 CATEGORY_CODE_MAP），從解析結果帶
    過來；抓不到就退回固定的 GUM糖，不要讓檔名整個開天窗。"""
    po = po_numbers[0] if po_numbers else ""
    if line_key == "pg":
        # 檔名不放單號——同一批匯出好幾張 PO 就會撞成同一個檔名，交給
        # 另一層檔名去重機制兜底（見 api_purchase_export，撞了會自動
        # 加 (2)、(3)，不會互相覆蓋，只是檔名看不出是哪一張）。
        return f"酷澎XP&G_產品採購表上傳_{date_mmdd}到貨.xls"
    if line_key == "paper":
        return f"酷澎_產品採購表上傳_{date_mmdd}到貨.xls"
    if line_key == "mars":
        wh = warehouse or "倉別"
        cat = category or "GUM糖"
        suffix = f"_{po}" if po else ""
        return f"永豐Mars採購單(箱單位)-{cat}_{wh}{suffix}.xls"
    return f"採購表_{date_mmdd}.xls"


def fill_template(line_key, rows, remark):
    """載入該線別的公司範本檔，清空資料區、寫入新資料，回傳檔案位元組。

    用 xlutils.copy 而不是重新用 xlwt 從零產生，是為了讓查表工作表
    （表單選項／商品資料／客編）跟儲存格樣式完全保留，降低跟公司系統
    要求的格式對不起來的風險。

    備註是「逐列」的：合併多張 PO 時，每一列要保留自己那張 PO 的備註
    （P&G／瑪氏的備註帶著訂單編號），所以每列各自帶 remark 就用自己的，
    沒帶才退回用整份共用的那個。之前整份共用一個備註，合併後會把所有
    列都寫成第一張 PO 的單號，後面幾張的單號整個消失。"""
    cfg = LINES[line_key]
    rb = xlrd.open_workbook(cfg["template"], formatting_info=True)
    wb = xl_copy(rb)
    ws = wb.get_sheet(0)

    wipe_upto = max(cfg["wipe_rows"], len(rows) + 1)
    for r in range(1, wipe_upto):
        for c in range(10):
            ws.write(r, c, "")

    for i, row in enumerate(rows, start=1):
        ws.write(i, 0, i)                        # 項目
        ws.write(i, 1, str(row["material_no"]))   # 料號
        ws.write(i, 2, "")                         # 品名
        ws.write(i, 3, "箱")                        # 單位
        ws.write(i, 6, int(row["qty"]))            # 採購數量
        ws.write(i, 8, row.get("remark", remark))   # 備註（逐列，見上面說明）

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------- 路由

@purchase_bp.route("/purchase")
def purchase_page():
    # 這頁的識別色是綠色，logo 也另外用一張綠色版。檔案還沒放進去之前
    # 先用原本那張，等 static/logo_purchase.png 一丟進去就自動換掉，
    # 不用再改程式；萬一哪天被刪掉也只是變回原本的 logo，不會破圖。
    logo = ("logo_purchase.png"
            if os.path.exists(os.path.join(os.path.dirname(__file__),
                                            "static", "logo_purchase.png"))
            else "logo.png")
    return render_template("purchase.html", lines=LINES, logo_file=logo)


@purchase_bp.route("/api/purchase/lines")
def api_purchase_lines():
    return jsonify([{"key": k, "label": v["label"]} for k, v in LINES.items()])


@purchase_bp.route("/api/purchase/warehouses")
def api_purchase_warehouses():
    return jsonify(_load_warehouses())


@purchase_bp.route("/api/purchase/parse", methods=["POST"])
def api_purchase_parse():
    line_key = request.form.get("line", "")
    if line_key not in LINES:
        return jsonify({"error": "請選擇正確的線別。"}), 400
    uploads = [f for f in request.files.getlist("file") if f and f.filename]
    if not uploads:
        return jsonify({"error": "沒有收到檔案。"}), 400

    if LINES[line_key].get("multi_file"):
        # 瑪氏一份訂貨通知單就是一張 PO，可以一次選多個檔案，各自轉成
        # 各自的一組。整批一起成功或一起失敗（跟系統其他地方的「全有
        # 全無」原則一樣）：只要有一個檔案解析失敗就整批擋下來、講清楚
        # 是哪個檔案有問題，不要讓使用者匯出一半、缺了某張單卻不知道。
        groups = []
        errors = []
        for upload in uploads:
            try:
                groups.append(parse_mars_order_form(upload.read(), upload.filename))
            except PurchaseImportError as exc:
                errors.append(f"「{upload.filename}」：{exc}")
        if errors:
            return jsonify({"error": "有檔案解析失敗，本次都不會匯入：\n" + "\n".join(errors)}), 400
    else:
        if len(uploads) > 1:
            return jsonify({
                "error": f"{LINES[line_key]['label']}一次只能上傳一個檔案（瑪氏才能一次選多個）。"}), 400
        try:
            groups = parse_import_file(line_key, uploads[0].read(), uploads[0].filename)
        except PurchaseImportError as exc:
            return jsonify({"error": str(exc)}), 400

    for g in groups:
        category = g.get("category", "")
        g["remark"] = build_remark(line_key, g["po_numbers"], g["date_guess"], g["warehouse_guess"])
        # 瑪氏是「一份訂貨通知單＝一張 PO＝一個檔」，匯出檔名直接沿用
        # 匯入檔名（見 parse_mars_order_form），不用猜出來的日期／倉別／
        # 品類拼；猜錯或漏猜的欄位不會連累到檔名。P&G／紙潔本來就可能
        # 多張 PO 合併成一份，沒有單一原檔名可以沿用，維持原本的規則。
        stem = g.pop("filename_stem", "")
        if line_key == "mars" and stem:
            g["filename"] = stem + ".xls"
        else:
            g["filename"] = build_filename(
                line_key, g["po_numbers"], g["date_guess"], g["warehouse_guess"], category)

    return jsonify({"line": line_key, "groups": groups})


@purchase_bp.route("/api/purchase/export", methods=["POST"])
def api_purchase_export():
    payload = request.get_json(silent=True) or {}
    line_key = payload.get("line", "")
    if line_key not in LINES:
        return jsonify({"error": "請選擇正確的線別。"}), 400
    groups = payload.get("groups") or []
    if not isinstance(groups, list) or not groups:
        return jsonify({"error": "沒有要匯出的資料。"}), 400

    files = []
    for g in groups:
        rows = g.get("rows") or []
        if not rows:
            continue
        filename = (g.get("filename") or "").strip() or "採購表.xls"
        if not filename.lower().endswith(".xls"):
            filename += ".xls"
        remark = g.get("remark") or ""
        data = fill_template(line_key, rows, remark)
        files.append((filename, data))

    if not files:
        return jsonify({"error": "選取的項目裡沒有任何資料列。"}), 400

    if len(files) == 1:
        name, data = files[0]
        return send_file(io.BytesIO(data), as_attachment=True, download_name=name,
                          mimetype="application/vnd.ms-excel")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        used_names = set()
        for name, data in files:
            final = name
            n = 2
            while final in used_names:
                base, ext = os.path.splitext(name)
                final = f"{base}({n}){ext}"
                n += 1
            used_names.add(final)
            zf.writestr(final, data)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="採購表.zip",
                      mimetype="application/zip")


@purchase_bp.route("/api/purchase/warehouses", methods=["POST"])
def api_purchase_warehouses_save():
    payload = request.get_json(silent=True) or {}
    address = (payload.get("address") or "").strip()
    code = (payload.get("code") or "").strip()
    if not address or not code:
        return jsonify({"error": "地址跟倉別代碼都要填。"}), 400
    mapping = _load_warehouses()
    mapping[address] = code
    _save_warehouses(mapping)
    return jsonify({"ok": True, "warehouses": mapping})
