"""酷澎訂單管理系統 — 本地端中繼站。

啟動：python app.py  然後開 http://127.0.0.1:5000
"""

import base64
import datetime as _dt
import io
import json
import os
import secrets
import zipfile
from functools import wraps

from flask import (
    Flask, jsonify, redirect, render_template, request, send_file, session,
    url_for,
)
from werkzeug.security import check_password_hash, generate_password_hash

import db
from db import get_conn, init_db, now
from importer import (
    COUPANG_FIELDS, CRITICAL_FIELDS, ImportError_, desired_ship, diff_rows,
    parse_workbook,
)
import pdfsign
from normalize import norm_date, norm_int, norm_key, norm_text, norm_warehouse

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 每次改版就手動往上加一號。畫面右下角跟啟動視窗都會印出這個號碼，
# 用來確認「現在看到的畫面」跟「最新給的檔案」是不是同一份——
# 之前吃過虧：舊的黑視窗沒關乾淨，背景還留著一個沒更新到的伺服器
# 在跑，怎麼換檔案畫面都不會變，肉眼完全看不出來是這個原因。
BUILD_VERSION = "2026-08-26.6"

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024
# 模板一律即時讀檔，不要用啟動當下快取的舊版本（debug=False 時 Flask
# 預設不會自動重讀模板檔，只改 index.html 卻要重開程式才生效，容易
# 造成「明明換了檔案，畫面卻沒變」的誤判）。
app.config["TEMPLATES_AUTO_RELOAD"] = True

# 只要「開了資料資料夾、建了資料庫表」這件事，不管是 python app.py
# 直接跑，還是 gunicorn 匯入 app 物件來跑，都要做——之前這兩行只寫
# 在最底下 `if __name__ == "__main__":` 裡，gunicorn 用 import 的方式
# 啟動時完全不會執行到那段，資料夾、資料庫表都不會被建出來，一部署
# 上 Render 就整個掛掉。搬到這裡，模組被 import 的當下就一定會跑到。
moved_on_start = db.ensure_data_dir()
init_db()


def _load_or_create_secret_key():
    # 沒設 SECRET_KEY 環境變數的話，把 session 簽章金鑰存進資料資料夾，
    # 一次產生、之後重複使用——不然每次重開程式（或 Render 重新部署）
    # 金鑰都換一把，所有人都會被登出。
    env_key = os.environ.get("SECRET_KEY")
    if env_key:
        return env_key
    path = os.path.join(db.DATA_DIR, ".secret_key")
    try:
        with open(path, encoding="utf-8") as fh:
            key = fh.read().strip()
            if key:
                return key
    except OSError:
        pass
    key = secrets.token_hex(32)
    try:
        os.makedirs(db.DATA_DIR, exist_ok=True)
        with open(path, "w", encoding="utf-8") as fh:
            fh.write(key)
    except OSError:
        pass
    return key


def _load_or_create_sync_token():
    """外部同步（酷澎後台驗收工具）用的通行碼，不是給人登入用的密碼，
    是給那支瀏覽器腳本證明「這個請求真的是我們授權的來源」用的。
    存在資料庫的 app_settings，不用檔案——Postgres 模式下（Render）
    重新部署，本機檔案會被換成全新的，token 跟著換一把，腳本那邊
    存的舊 token 就失效了，同事還得回來重貼一次。存進資料庫才會
    跟訂單資料一樣，重新部署也不會變。"""
    env_val = os.environ.get("SYNC_TOKEN")
    if env_val:
        return env_val
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT value FROM app_settings WHERE key = ?", ("sync_token",)).fetchone()
        if row is not None:
            return json.loads(row["value"])
    finally:
        conn.close()
    token = secrets.token_hex(20)
    db.save_setting("sync_token", token)
    return token


app.secret_key = _load_or_create_secret_key()
# 部署在外網時 Render 是走 HTTPS，這樣 cookie 只會經加密連線傳送；
# 本機用 http://127.0.0.1 開發一樣有效，Flask 不會因此擋掉本機測試。
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"

# 單一 SKU 自己的欄位，改一個品項不影響同張單的其他品項。
# 匯入永遠不會覆蓋這些欄位。備註已搬到整張 PO 共用（見 PO_EDITABLE_FIELDS），
# 不再是品項各自一份。
EDITABLE_FIELDS = {
    "qty_ship":       "出貨數量",
    "receiving_note": "驗收註記",
}

# 改這個欄位時，順便標記「人工調整過」，下次匯入酷澎的檔案就不再覆蓋
# ——跟交期／倉別是同一套邏輯，只是這是每筆 SKU 各自的值，不是整張
# 單一起改。箱入數目前是唯讀欄位，不開放編輯，不在這個清單裡。
SKU_OVERRIDE_FIELDS = ("qty_ship",)

# 拉單鎖定要擋的是「會影響倉庫出貨的數字」；驗收註記是事後才填的資訊
# （短驗、溢收…都是貨到了才知道），鎖住反而擋住正常作業，所以不受拉單
# 鎖定限制，任何時候都能改。
ALWAYS_EDITABLE_FIELDS = {"receiving_note"}

# 整張 PO 共用的欄位，改了就是整張單一起改（OP 拋 ERP、交倉庫都是整張
# 單一起行動，不會拆開）。存在 po_headers，匯入一律不覆蓋。
PO_EDITABLE_FIELDS = {
    "po_status":        "PO狀態",
    "receiving_status": "驗收狀態",
    "filed_date":       "建檔日",
    "shipping_method":  "配送方式",
    "remarks":          "備註",
}

# 驗收狀態不受拉單鎖定限制——驗收本來就發生在拉單之後（貨到了才驗），
# 鎖住它會直接擋住正常作業流程，跟備註／驗收註記不受鎖定是同一個道理。
# 配送方式是匯入後才由 OP 自己補的內部分類，跟出貨數量無關，一樣不該
# 被拉單鎖住。備註本來就是整張單事後才寫的資訊，搬到 PO 層級後同樣
# 不受拉單鎖定影響。
PO_ALWAYS_EDITABLE_FIELDS = {"receiving_status", "shipping_method", "remarks"}

# 整張 PO 共用、但屬於酷澎來源的欄位：OP 可以改（跟酷澎談好調整），
# 改的時候整張單一起改，但它們存在 orders 上、且會參與匯入比對。
PO_COUPANG_FIELDS = {
    "delivery_date": "交期",
    "warehouse":     "倉別",
    "order_type":    "訂單類型",
}

# PO_COUPANG_FIELDS 各自要用的正規化函式——倉別要轉大寫、交期要轉
# ISO 日期，訂單類型只是自由文字，直接照打的存。
_PO_COUPANG_FIELD_NORM = {
    "delivery_date": norm_date,
    "warehouse":     norm_warehouse,
    "order_type":    norm_text,
}

# 酷澎來源欄位中，允許 OP「補空白」的欄位。
# 只有原本是空白時才能填，已經有值的一律不給改——避免有人手滑把瑪氏
# 改成寶僑。上游主檔之後補建了，匯入仍會以主檔為準覆蓋回來並記歷程。
FILLABLE_FIELDS = {
    "line":  "線別",
    "brand": "品牌",
}

INSERT_FIELDS = [
    "po_number", "sku_id", "order_type", "parent_po", "line", "brand",
    "product_name", "barcode", "yf_sku", "warehouse", "address",
    "delivery_date", "qty_coupang", "unit", "box_size", "unit_price",
    "expiry_note", "seq_no",
]


# ---------------------------------------------------------------- 設定檔

# PostgreSQL 模式下這三個設定檔改存進資料庫（見 db.load_setting／
# save_setting）；SQLite 模式維持原本放檔案的做法。
_SETTINGS_KEY = {
    "config.json": "config",
    "users.json": "users",
    "export_profiles.json": "export_profiles",
}


def load_json(name, default):
    if db.IS_POSTGRES:
        return db.load_setting(_SETTINGS_KEY[name], default)
    # 設定檔住在資料資料夾，不在程式資料夾——更新版本時不會被覆蓋
    path = os.path.join(db.DATA_DIR, name)
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, ValueError):
        return default


def get_config():
    return load_json("config.json", {
        "operators": ["OP"],
        "po_statuses": ["已建立", "已回覆", "處理中", "已完成", "修改中", "已取消"],
        "receiving_statuses": ["未驗收", "完成", "異常", "重啟", "退貨"],
        "order_types": ["一般", "NS", "補單", "拆單"],
        "shipping_methods": ["原廠(EM)", "竹運(CUP)"],
    })


def get_profiles():
    return load_json("export_profiles.json", {"profiles": {}}).get("profiles", {})


def get_users():
    """帳號密碼放 users.json，跟 config.json 一樣住在資料資料夾，改完
    存檔、使用者重新登入就生效，不用改程式、不用重新部署。"""
    raw = load_json("users.json", {})
    # 檔案裡的 "_說明" 那行是給人看的註解，不是帳號，過濾掉——不然
    # 有人把使用者名稱打成「_說明」，密碼打那串說明文字，就真的能登入。
    return {k: v for k, v in raw.items() if not k.startswith("_")}


def save_json(name, data):
    """寫設定檔。先寫暫存檔再換過去，中途斷電也不會留下半個壞掉的
    JSON——設定檔壞掉會讓所有人登不進來，這個險不值得冒。"""
    if db.IS_POSTGRES:
        db.save_setting(_SETTINGS_KEY[name], data)
        return
    path = os.path.join(db.DATA_DIR, name)
    tmp = path + ".tmp"
    os.makedirs(db.DATA_DIR, exist_ok=True)
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _ensure_receiving_status_option():
    """新增「退貨」這個驗收狀態選項，補進既有安裝裡。

    PostgreSQL 模式下設定值已經存進資料庫，只有第一次啟動、資料庫裡
    完全沒有這筆設定時，defaults/config.json 才會被種進去——已經跑
    過一陣子的正式站不會再吃到新的預設檔，所以這裡要用一次性補寫，
    只在清單裡沒有「退貨」時才加，不動使用者自己調整過的其他選項。"""
    cfg = get_config()
    statuses = list(cfg.get("receiving_statuses") or [])
    if "退貨" not in statuses:
        statuses.append("退貨")
        cfg["receiving_statuses"] = statuses
        save_json("config.json", cfg)


def _ensure_shipping_methods_option():
    """新增「配送方式」選項清單，補進既有安裝裡——道理跟補「退貨」
    驗收狀態一樣：正式站的設定值已經存進資料庫，只改 defaults 檔案
    對已經跑起來的安裝沒有用。"""
    cfg = get_config()
    if not cfg.get("shipping_methods"):
        cfg["shipping_methods"] = ["原廠(EM)", "竹運(CUP)"]
        save_json("config.json", cfg)


def _ensure_export_profile_columns():
    """把「配送方式」「個人標記」兩個新欄位補進既有安裝的「完整欄位」
    匯出格式裡——道理跟上面兩個 _ensure_* 一樣：正式站的匯出格式已經
    存進資料庫，只改 export_profiles.json 對已經跑起來的安裝沒有用。
    只補「完整欄位」這個對帳／備份用的格式，不去動倉庫出貨表、ERP
    格式——那兩個是給外部/工廠看的，OP 自己的顏色標記跟他們無關；
    也不動 OP 自己在「完整欄位」裡已經調整過的欄位順序或刪掉的欄位，
    只在確定沒有這個欄位代碼時才補加到最後面。"""
    data = load_json("export_profiles.json", {"profiles": {}})
    full = (data.get("profiles") or {}).get("full")
    if not full:
        return
    columns = full.get("columns") or []
    existing_fields = {c.get("field") for c in columns}
    changed = False
    if "shipping_method" not in existing_fields:
        columns.append({"header": "配送方式", "field": "shipping_method"})
        changed = True
    if "flagged" not in existing_fields:
        columns.append({"header": "個人標記", "field": "flagged", "format": "yesno"})
        changed = True
    if changed:
        full["columns"] = columns
        save_json("export_profiles.json", data)


_ensure_receiving_status_option()
_ensure_shipping_methods_option()
_ensure_export_profile_columns()


def verify_password(stored, given):
    """比對密碼，同時支援明碼與 werkzeug 雜湊。

    不要寫死判斷 "pbkdf2:" 前綴：werkzeug 換過預設演算法（pbkdf2 →
    scrypt），寫死前綴會讓新設的密碼被當成明碼去比對，於是「在設定頁
    改完密碼 → 再也登不進來」，而且錯得無聲無息。改看有沒有 `$`——
    werkzeug 的雜湊一律是 method$salt$hash，人手打的明碼不會長這樣。"""
    if not stored:
        return False
    if "$" in stored:
        try:
            return check_password_hash(stored, given)
        except (ValueError, TypeError):
            return False
    return secrets.compare_digest(stored, given)


def current_operator():
    """操作人員一律取登入者，不再由前端傳。

    以前是右上角自己挑名字，等於誰都能把自己做的修改掛到別人頭上，
    稽核歷程就失去意義了；有了登入之後，session 裡的身分才是唯一
    可信的來源。所有 API 都在 require_login 後面，這裡必有值。"""
    return session.get("user", "")


# ---------------------------------------------------------------- 登入
# 這系統本來只跑在辦公室內網，沒有密碼也還好；一旦放到外網，網址誰都
# 能打開、誰都能看到全部訂單、誰都能改，所以只要走外網就一定要有登入。

PUBLIC_ENDPOINTS = {"login", "static"}


@app.before_request
def require_login():
    if request.endpoint in PUBLIC_ENDPOINTS or request.endpoint is None:
        return
    if session.get("user"):
        return
    if request.path.startswith("/api/"):
        return jsonify({"error": "未登入，請重新整理頁面登入"}), 401
    return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        if session.get("user"):
            return redirect(url_for("index"))
        return render_template("login.html", error=None, build_version=BUILD_VERSION)

    username = (request.form.get("username") or "").strip()
    password = request.form.get("password") or ""
    users = get_users()
    # 支援兩種寫法：一開始給的是明碼方便你自己改，透過設定頁改過的
    # 就是雜湊值，verify_password 兩種都認得。
    if not verify_password(users.get(username, ""), password):
        return render_template(
            "login.html", error="帳號或密碼不對，再檢查一次",
            build_version=BUILD_VERSION,
        ), 401

    session.clear()
    session["user"] = username
    session.permanent = True
    nxt = request.args.get("next") or request.form.get("next") or "/"
    if not nxt.startswith("/"):
        nxt = "/"
    return redirect(nxt)


@app.route("/logout", methods=["POST"])
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------------- 帳號管理
# 部署到 Render 之後就沒有「用記事本打開設定檔」這個選項了——雲端機器
# 上沒有檔案總管。加人、改密碼、清資料這三件事只能在系統裡面做完，
# 否則密碼永遠停在預設值。

def get_admins():
    """管理員名單放 users.json 的 `_admins`。底線開頭的鍵本來就會被
    get_users() 濾掉，不會被誤認成帳號，剛好拿來放這種設定。

    名單空的時候一律當成「所有人都是管理員」——不然升級上來的舊檔案
    沒有這個欄位，就會變成沒有人能進設定頁，自己把自己鎖在門外。"""
    raw = load_json("users.json", {})
    admins = raw.get("_admins")
    if isinstance(admins, list) and admins:
        return {str(a) for a in admins}
    return set(get_users().keys())


def is_admin(name=None):
    return (name or current_operator()) in get_admins()


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not is_admin():
            return jsonify({"error": "只有管理員可以做這個操作。"}), 403
        return fn(*args, **kwargs)
    return wrapper


def _write_users(users, admins):
    save_json("users.json", {
        "_說明": ("登入帳號密碼。請用系統裡的「設定」頁面修改，"
                  "手動改這個檔案也可以，密碼支援明碼或 pbkdf2 雜湊。"),
        "_admins": sorted(admins),
        **users,
    })


@app.route("/api/account")
def api_account():
    """畫面要知道「我是誰、我是不是管理員」才決定設定頁顯示到哪。"""
    me = current_operator()
    return jsonify({
        "user": me,
        "is_admin": is_admin(me),
        "users": sorted(get_users().keys()) if is_admin(me) else [],
        "admins": sorted(get_admins()) if is_admin(me) else [],
    })


@app.route("/api/account/password", methods=["POST"])
def api_change_password():
    """改密碼。改自己的要驗舊密碼；管理員可以直接重設別人的（有人忘記
    密碼時總要有人能救），但不必也不能藉此看到對方原本的密碼。"""
    payload = request.get_json(silent=True) or {}
    me = current_operator()
    target = norm_text(payload.get("username")) or me
    new_password = payload.get("new_password") or ""
    old_password = payload.get("old_password") or ""

    if len(new_password) < 6:
        return jsonify({"error": "新密碼至少 6 個字。"}), 400

    users = get_users()
    if target not in users:
        return jsonify({"error": f"找不到帳號「{target}」。"}), 404

    if target != me:
        if not is_admin(me):
            return jsonify({"error": "只有管理員可以重設別人的密碼。"}), 403
    elif not verify_password(users[target], old_password):
        return jsonify({"error": "舊密碼不對。"}), 400

    users[target] = generate_password_hash(new_password)
    _write_users(users, get_admins())
    return jsonify({"ok": True, "message": f"「{target}」的密碼已更新。"})


@app.route("/api/account/users", methods=["POST"])
@admin_required
def api_add_user():
    payload = request.get_json(silent=True) or {}
    name = norm_text(payload.get("username"))
    password = payload.get("password") or ""
    make_admin = bool(payload.get("is_admin"))

    if not name:
        return jsonify({"error": "請填帳號名稱。"}), 400
    if name.startswith("_"):
        return jsonify({"error": "帳號名稱不能用底線開頭。"}), 400
    if len(password) < 6:
        return jsonify({"error": "密碼至少 6 個字。"}), 400

    users = get_users()
    if name in users:
        return jsonify({"error": f"帳號「{name}」已經存在。"}), 400

    users[name] = generate_password_hash(password)
    admins = get_admins()
    if make_admin:
        admins.add(name)
    _write_users(users, admins)
    return jsonify({"ok": True, "message": f"已新增帳號「{name}」。"})


@app.route("/api/account/users/delete", methods=["POST"])
@admin_required
def api_delete_user():
    payload = request.get_json(silent=True) or {}
    name = norm_text(payload.get("username"))
    users = get_users()

    if name not in users:
        return jsonify({"error": f"找不到帳號「{name}」。"}), 404
    if name == current_operator():
        return jsonify({"error": "不能刪除自己的帳號。"}), 400
    if len(users) <= 1:
        return jsonify({"error": "至少要保留一個帳號。"}), 400

    admins = get_admins() - {name}
    if not admins:
        return jsonify({"error": "刪掉他就沒有管理員了，請先指定其他管理員。"}), 400

    del users[name]
    _write_users(users, admins & set(users.keys()))
    # 帳號刪掉，但 edit_logs 裡他做過的修改一律保留——歷程是對帳用的
    # 證據，不能因為人離職就消失。
    return jsonify({"ok": True, "message": f"已刪除帳號「{name}」。"})


@app.route("/api/account/admin", methods=["POST"])
@admin_required
def api_set_admin():
    payload = request.get_json(silent=True) or {}
    name = norm_text(payload.get("username"))
    make_admin = bool(payload.get("is_admin"))
    users = get_users()

    if name not in users:
        return jsonify({"error": f"找不到帳號「{name}」。"}), 404

    admins = get_admins()
    if make_admin:
        admins.add(name)
    else:
        admins.discard(name)
        if not admins:
            return jsonify({"error": "至少要保留一個管理員。"}), 400

    _write_users(users, admins)
    word = "設為管理員" if make_admin else "取消管理員"
    return jsonify({"ok": True, "message": f"已將「{name}」{word}。"})


@app.route("/api/account/reset-data", methods=["POST"])
@admin_required
def api_reset_data():
    """清空訂單資料。帳號、設定、修改歷程要不要一起清，分開讓人選——
    最常見的情況是「測試資料清掉、正式開始用」，那時歷程也該一起清；
    但也有「只想重來一次匯入」的情況，歷程留著才查得到之前發生什麼。"""
    payload = request.get_json(silent=True) or {}
    if norm_text(payload.get("confirm")) != "清空資料":
        return jsonify({"error": "請照著輸入「清空資料」四個字再確認。"}), 400

    keep_logs = bool(payload.get("keep_logs"))
    backup = db.backup_db("reset")

    conn = get_conn()
    try:
        # 先刪子表再刪主表，順序反了會踩到外鍵
        conn.execute("DELETE FROM export_batch_items")
        conn.execute("DELETE FROM export_batches")
        conn.execute("DELETE FROM import_batches")
        conn.execute("DELETE FROM orders")
        conn.execute("DELETE FROM po_headers")
        if not keep_logs:
            conn.execute("DELETE FROM edit_logs")
        conn.commit()
    finally:
        conn.close()

    tail = "，修改歷程保留" if keep_logs else "，修改歷程一併清除"
    note = f"（清空前已自動備份：{os.path.basename(backup)}）" if backup else ""
    return jsonify({"ok": True, "message": f"訂單資料已清空{tail}。{note}"})


@app.route("/api/account/sync-token")
@admin_required
def api_get_sync_token():
    return jsonify({"token": _load_or_create_sync_token()})


@app.route("/api/account/sync-token/regenerate", methods=["POST"])
@admin_required
def api_regenerate_sync_token():
    """換一把新的，舊的立刻失效——外部工具那邊要記得跟著改，不然
    會開始被 401 擋下來。"""
    token = secrets.token_hex(20)
    db.save_setting("sync_token", token)
    return jsonify({"ok": True, "token": token})


# ---------------------------------------------------------------- 驗收單簽名
#
# 取代原本要開 Google Colab 貼程式碼跑的作業（SOP-CP-CPG-001）。做的事
# 情一樣：在驗收單 PDF 上找「出貨確認（廠商簽名）」，蓋上簽名圖。
#
# 簽名圖一人一張、存在系統裡：登入誰蓋的就是誰的章，跟修改歷程記登入
# 身分同一套精神——共用一張圖的話，事後查不出這份驗收單是誰簽的。

# PDF 本體佔空間，資料庫容量有限，預設留半年就清掉；但「誰在什麼時候
# 簽了哪張單」的紀錄永久保留，不跟著被清。
SIGN_DEFAULT_RETENTION_DAYS = 180
MAX_SIGNATURE_BYTES = 2 * 1024 * 1024
MAX_SIGN_PDF_BYTES = 25 * 1024 * 1024


def get_sign_settings():
    """除了尺寸／位移／關鍵字／保留天數，也可能存著校正精靈上次用過
    的範例 PDF（calibration_pdf_b64／calibration_filename）——這裡直接
    整包合併回傳，不逐欄位篩選，不然這兩個欄位會在讀回來時被濾掉。"""
    saved = db.load_setting("sign_settings", {}) if db.IS_POSTGRES else \
        load_json("config.json", {}).get("sign_settings", {})
    geo = dict(pdfsign.DEFAULT_GEOMETRY)
    geo["retention_days"] = SIGN_DEFAULT_RETENTION_DAYS
    if isinstance(saved, dict):
        geo.update({k: v for k, v in saved.items() if v not in (None, "")})
    return geo


def save_sign_settings(geo):
    if db.IS_POSTGRES:
        db.save_setting("sign_settings", geo)
        return
    cfg = load_json("config.json", {})
    cfg["sign_settings"] = geo
    save_json("config.json", cfg)


@app.route("/api/sign/settings")
def api_sign_settings():
    geo = get_sign_settings()
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT filename, byte_size, updated_at FROM user_signatures "
            "WHERE username = ?", (current_operator(),)).fetchone()
    finally:
        conn.close()
    return jsonify({
        "settings": geo,
        "signature": dict(row) if row else None,
        "is_admin": is_admin(),
    })


@app.route("/api/sign/settings", methods=["POST"])
def api_save_sign_settings():
    """尺寸／位移／關鍵字改成設定值，不寫死在程式裡——酷澎哪天改了
    版面或字，或是簽名蓋歪了要微調，任何人都能自己改，不用等改程式、
    重新部署。原本 SOP 的異常處理第 4 點就是教人去改程式碼裡那四個
    數字，那對非技術同事等於做不到。

    這是全公司共用一份的設定，開放所有登入的人都能調，不限管理員——
    改了會影響所有人的簽名位置，但校正介面是所見即所得（拖曳當場就
    看得到蓋在哪），改錯了重拖一次就好，不是那種盲改。唯一的保險是
    每次存檔都記下是誰、什麼時候改的，畫面上看得到，出問題不用瞎猜。"""
    payload = request.get_json(silent=True) or {}
    geo = get_sign_settings()

    # 關鍵字只能去頭尾空白，不能走 norm_text——它會做 NFKC 正規化，把
    # 全形括號「（）」轉成半形「()」。PDF 裡印的是全形，轉過就再也對不
    # 上，結果是每份檔案都說「找不到簽名欄位」，而且看起來完全正常，
    # 極難查。這裡要的是跟 PDF 內文逐字相符，不是好看的正規化結果。
    keyword = str(payload.get("keyword") or "").strip()
    if not keyword:
        return jsonify({"error": "關鍵字不能空白。"}), 400
    geo["keyword"] = keyword

    for key in ("width", "height", "offset_x", "offset_y"):
        if key not in payload:
            continue
        value = norm_int(payload.get(key))
        if value is None:
            return jsonify({"error": f"{key} 請填數字。"}), 400
        if key in ("width", "height") and value <= 0:
            return jsonify({"error": "簽名寬高要大於 0。"}), 400
        geo[key] = value

    if "retention_days" in payload:
        days = norm_int(payload.get("retention_days"))
        if days is None or days < 0:
            return jsonify({"error": "保留天數請填 0 以上的數字。"}), 400
        geo["retention_days"] = days

    geo["updated_by"] = current_operator()
    geo["updated_at"] = now()
    save_sign_settings(geo)
    return jsonify({"ok": True, "settings": geo})


@app.route("/api/sign/calibrate/upload", methods=["POST"])
def api_sign_calibrate_upload():
    """上傳一份真的驗收單，找到關鍵字所在頁渲染成圖，讓人用拖曳的
    方式校正簽名要蓋在哪——取代原本要填四個數字用猜的。這份範例
    PDF 會存起來，下次要重新校正不用再傳一次（見 current 那支）。"""
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "沒有收到 PDF 檔。"}), 400
    raw = upload.read()
    if len(raw) > MAX_SIGN_PDF_BYTES:
        return jsonify({"error": "檔案太大（超過 25MB）。"}), 400

    keyword = str(request.form.get("keyword") or "").strip() \
        or get_sign_settings()["keyword"]
    try:
        info = pdfsign.render_for_calibration(raw, keyword)
    except pdfsign.SignError as exc:
        return jsonify({"error": str(exc)}), 400

    geo = get_sign_settings()
    geo["calibration_pdf_b64"] = base64.b64encode(raw).decode("ascii")
    geo["calibration_filename"] = upload.filename
    save_sign_settings(geo)

    return jsonify({
        "ok": True, "keyword": keyword,
        "image_b64": base64.b64encode(info["image_bytes"]).decode("ascii"),
        "dpi": info["dpi"], "image_width": info["image_width"],
        "image_height": info["image_height"],
        "keyword_rect_px": info["keyword_rect_px"],
    })


@app.route("/api/sign/calibrate/current")
def api_sign_calibrate_current():
    """重新打開校正畫面時，用上次存的範例 PDF 重新渲染一次，不用
    使用者再傳一次檔案。"""
    geo = get_sign_settings()
    pdf_b64 = geo.get("calibration_pdf_b64")
    if not pdf_b64:
        return jsonify({"error": "還沒有校正過的範例 PDF。"}), 404

    keyword = str(request.args.get("keyword") or "").strip() or geo["keyword"]
    try:
        info = pdfsign.render_for_calibration(base64.b64decode(pdf_b64), keyword)
    except pdfsign.SignError as exc:
        return jsonify({"error": str(exc)}), 400

    return jsonify({
        "ok": True, "keyword": keyword, "filename": geo.get("calibration_filename", ""),
        "image_b64": base64.b64encode(info["image_bytes"]).decode("ascii"),
        "dpi": info["dpi"], "image_width": info["image_width"],
        "image_height": info["image_height"],
        "keyword_rect_px": info["keyword_rect_px"],
    })


@app.route("/api/sign/signature", methods=["POST"])
def api_upload_signature():
    """上傳自己的簽名圖。存起來之後就不用每次簽名都重傳，也就沒有
    原本 Colab 版「一次只能傳一個簽名檔」那條要人自己記得的規則。"""
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "沒有收到圖片檔。"}), 400

    raw = upload.read()
    if not raw:
        return jsonify({"error": "檔案是空的。"}), 400
    if len(raw) > MAX_SIGNATURE_BYTES:
        return jsonify({"error": "簽名圖請小於 2MB。"}), 400

    try:
        info = pdfsign.validate_signature(raw)
    except pdfsign.SignError as exc:
        return jsonify({"error": str(exc)}), 400

    me = current_operator()
    stamp = now()
    conn = get_conn()
    try:
        conn.execute("DELETE FROM user_signatures WHERE username = ?", (me,))
        conn.execute(
            """INSERT INTO user_signatures
               (username, image_b64, mime, filename, byte_size, updated_at)
               VALUES (?,?,?,?,?,?)""",
            (me, base64.b64encode(raw).decode("ascii"),
             upload.mimetype or "image/png", upload.filename, len(raw), stamp))
        conn.commit()
    finally:
        conn.close()

    return jsonify({"ok": True, "message": "簽名圖已儲存。",
                    "filename": upload.filename, "byte_size": len(raw),
                    "updated_at": stamp, **info})


@app.route("/api/sign/signature", methods=["DELETE"])
def api_delete_signature():
    conn = get_conn()
    try:
        conn.execute("DELETE FROM user_signatures WHERE username = ?",
                     (current_operator(),))
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "message": "簽名圖已移除。"})


@app.route("/api/sign/signature/preview")
def api_preview_signature():
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT image_b64, mime FROM user_signatures WHERE username = ?",
            (current_operator(),)).fetchone()
    finally:
        conn.close()
    if row is None:
        return jsonify({"error": "還沒上傳簽名圖。"}), 404
    return send_file(io.BytesIO(base64.b64decode(row["image_b64"])),
                     mimetype=row["mime"] or "image/png")


@app.route("/api/sign/run", methods=["POST"])
def api_sign_run():
    """一次簽一批。每份檔案各自回報結果，某一份失敗不影響其他份——
    Colab 版遇到壞檔會整個中斷，前面簽好的也一起沒了。"""
    me = current_operator()
    uploads = request.files.getlist("files")
    uploads = [u for u in uploads if u and u.filename]
    if not uploads:
        return jsonify({"error": "沒有收到任何 PDF。"}), 400

    conn = get_conn()
    try:
        sig_row = conn.execute(
            "SELECT image_b64 FROM user_signatures WHERE username = ?",
            (me,)).fetchone()
    finally:
        conn.close()
    if sig_row is None:
        return jsonify({
            "error": "你還沒上傳簽名圖，請先到「設定 → 驗收單簽名」上傳一次。",
        }), 400
    signature = base64.b64decode(sig_row["image_b64"])

    geo = get_sign_settings()
    stamp = now()
    results = []

    conn = get_conn()
    try:
        cur = conn.execute(
            """INSERT INTO sign_batches
               (operator, file_count, signed_count, fail_count, keyword, created_at)
               VALUES (?,?,?,?,?,?)""",
            (me, len(uploads), 0, 0, geo["keyword"], stamp))
        batch_id = cur.lastrowid

        signed_total = 0
        fail_total = 0

        def record_failure(name, message):
            """失敗的也要進歸檔——「這份當初傳了但沒簽成功」跟「這份
            根本沒傳過」是兩件事，事後追查時差很多。"""
            po = pdfsign.extract_po_number("", name)
            conn.execute(
                """INSERT INTO signed_docs
                   (batch_id, operator, po_number, filename, sign_count,
                    status, message, pdf_b64, byte_size, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (batch_id, me, po, name, 0, "error", message, "", 0, stamp))
            results.append({"filename": name, "status": "error",
                            "message": message, "sign_count": 0,
                            "po_number": po})

        for upload in uploads:
            raw = upload.read()
            name = upload.filename
            if not name.lower().endswith(".pdf"):
                record_failure(name, "不是 PDF 檔，已略過。")
                fail_total += 1
                continue
            if len(raw) > MAX_SIGN_PDF_BYTES:
                record_failure(name, "檔案太大（超過 25MB）。")
                fail_total += 1
                continue

            try:
                out, count, po_number = pdfsign.sign_pdf(raw, signature, geo)
            except pdfsign.SignError as exc:
                record_failure(name, str(exc))
                fail_total += 1
                continue

            doc_cur = conn.execute(
                """INSERT INTO signed_docs
                   (batch_id, operator, po_number, filename, sign_count,
                    status, message, pdf_b64, byte_size, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (batch_id, me, po_number, name, count, "ok", "",
                 base64.b64encode(out).decode("ascii"), len(out), stamp))
            results.append({
                "filename": name, "status": "ok", "sign_count": count,
                "po_number": po_number, "doc_id": doc_cur.lastrowid,
                "byte_size": len(out),
            })
            signed_total += count

        conn.execute(
            "UPDATE sign_batches SET signed_count = ?, fail_count = ? WHERE id = ?",
            (signed_total, fail_total, batch_id))
        conn.commit()
    finally:
        conn.close()

    ok_count = sum(1 for r in results if r["status"] == "ok")
    _purge_old_signed_pdfs()
    return jsonify({
        "ok": True, "batch_id": batch_id, "results": results,
        "ok_count": ok_count, "fail_count": fail_total,
        "signed_total": signed_total,
        "message": (f"{ok_count} 份簽好了（共蓋 {signed_total} 處）"
                    + (f"，{fail_total} 份沒處理成功" if fail_total else "")),
    })


@app.route("/api/sign/batches/<int:batch_id>/download")
def api_sign_download_zip(batch_id):
    conn = get_conn()
    try:
        rows = conn.execute(
            """SELECT filename, pdf_b64 FROM signed_docs
               WHERE batch_id = ? AND status = 'ok' AND purged = 0
               ORDER BY id""", (batch_id,)).fetchall()
    finally:
        conn.close()
    if not rows:
        return jsonify({"error": "這個批次沒有可下載的檔案（可能已超過保留期限）。"}), 404

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for row in rows:
            zf.writestr(f"signed_{row['filename']}",
                        base64.b64decode(row["pdf_b64"]))
    buf.seek(0)
    return send_file(buf, mimetype="application/zip", as_attachment=True,
                     download_name=f"已簽名驗收單_{db.file_stamp()}.zip")


@app.route("/api/sign/docs/<int:doc_id>/download")
def api_sign_download_one(doc_id):
    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT filename, pdf_b64, purged FROM signed_docs WHERE id = ?",
            (doc_id,)).fetchone()
    finally:
        conn.close()
    if row is None or row["purged"] or not row["pdf_b64"]:
        return jsonify({"error": "這份檔案不存在或已超過保留期限被清除。"}), 404
    return send_file(io.BytesIO(base64.b64decode(row["pdf_b64"])),
                     mimetype="application/pdf", as_attachment=True,
                     download_name=f"signed_{row['filename']}")


@app.route("/api/sign/history")
def api_sign_history():
    """歸檔查詢：誰在什麼時候簽了哪張單。可依 PO 單號、操作人員篩。"""
    args = request.args
    where, params = [], []
    po = norm_text(args.get("po_number"))
    if po:
        where.append("po_number LIKE ?")
        params.append(f"%{po}%")
    operator = norm_text(args.get("operator"))
    if operator:
        where.append("operator = ?")
        params.append(operator)
    clause = f"WHERE {' AND '.join(where)}" if where else ""

    page_size = min(max(norm_int(args.get("page_size")) or 50, 1), 200)
    page = max(norm_int(args.get("page")) or 1, 1)

    conn = get_conn()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) AS n FROM signed_docs {clause}",
            params).fetchone()["n"]
        rows = conn.execute(
            f"""SELECT id, batch_id, operator, po_number, filename, sign_count,
                       status, message, byte_size, purged, created_at
                FROM signed_docs {clause}
                ORDER BY id DESC LIMIT ? OFFSET ?""",
            params + [page_size, (page - 1) * page_size]).fetchall()
        used = conn.execute(
            "SELECT COALESCE(SUM(byte_size), 0) AS n FROM signed_docs "
            "WHERE purged = 0").fetchone()["n"]
    finally:
        conn.close()

    return jsonify({
        "rows": [dict(r) for r in rows], "total": total,
        "page": page, "page_size": page_size,
        "stored_bytes": used,
        "retention_days": get_sign_settings()["retention_days"],
    })


@app.route("/api/sign/history/delete", methods=["POST"])
def api_sign_history_delete():
    """整筆刪掉，連紀錄一起消失——開放所有登入的人操作，跟位置校正
    同一個道理：公司就這幾個人，不是對外開放，做錯了也不是不可逆的
    資料損失（訂單資料才是）。跟「清除檔案」不一樣：那個只清 PDF
    本體，這個是真的整筆刪除。"""
    payload = request.get_json(silent=True) or {}
    ids = [i for i in (norm_int(x) for x in payload.get("doc_ids", [])) if i]
    if not ids:
        return jsonify({"error": "沒有選取任何紀錄。"}), 400

    conn = get_conn()
    try:
        placeholders = ",".join("?" * len(ids))
        conn.execute(f"DELETE FROM signed_docs WHERE id IN ({placeholders})", ids)
        conn.commit()
    finally:
        conn.close()
    return jsonify({"ok": True, "message": f"已刪除 {len(ids)} 筆紀錄。"})


@app.route("/api/sign/history/purge", methods=["POST"])
@admin_required
def api_sign_history_purge():
    """手動清除超過指定天數的 PDF 檔案本體，紀錄留著。跟自動清除
    （用設定頁的保留天數）共用同一個函式，差別只在天數哪裡來的。"""
    payload = request.get_json(silent=True) or {}
    days = norm_int(payload.get("days"))
    if days is None or days < 0:
        return jsonify({"error": "天數請填 0 以上的數字。"}), 400
    count = _purge_old_signed_pdfs(days)
    return jsonify({"ok": True, "message": f"已清除 {count} 份超過 {days} 天的檔案。"})


def _purge_old_signed_pdfs(days=None):
    """清掉超過保留天數的 PDF 本體，但保留「誰簽了什麼」的紀錄。

    保留天數設 0 表示不自動清。清的是 pdf_b64 這個大欄位，signed_docs
    那一列本身留著並標記 purged=1——歸檔查詢查得到這件事發生過，只是
    檔案本體不在了。回傳實際清了幾筆，手動清除時要回報給使用者看。
    """
    if days is None:
        days = get_sign_settings()["retention_days"]
    if not days:
        return 0
    cutoff = (db._local_now() - _dt.timedelta(days=days)).strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    try:
        cur = conn.execute(
            """UPDATE signed_docs SET pdf_b64 = '', byte_size = 0, purged = 1
               WHERE purged = 0 AND pdf_b64 != '' AND created_at < ?""",
            (cutoff,))
        count = cur.rowcount
        conn.commit()
        return count
    finally:
        conn.close()


# ---------------------------------------------------------------- 外部同步：
# 酷澎後台驗收工具（瀏覽器腳本）主動把「實際驗入數量」推進來，取代
# 「匯出 Excel 再手動上傳」那一步。跟一般 API 不一樣，呼叫的人不是
# 登入的使用者、是一支跑在酷澎網域裡的腳本，所以不能用 session 驗證，
# 改用一把只有這支腳本知道的 token；也因為呼叫來源是不同網域
# （supplier.tw.coupang.com），瀏覽器會先送一個 OPTIONS 預檢請求，
# 要主動回應 CORS 標頭，不然瀏覽器直接擋掉，連 POST 都不會送出。

SYNC_ALLOWED_ORIGIN = "https://supplier.tw.coupang.com"
PUBLIC_ENDPOINTS.add("api_sync_verified_qty")


@app.after_request
def _cors_for_sync(resp):
    if request.path.startswith("/api/sync/"):
        resp.headers["Access-Control-Allow-Origin"] = SYNC_ALLOWED_ORIGIN
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Sync-Token"
        resp.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS"
    return resp


@app.route("/api/sync/verified-qty", methods=["POST", "OPTIONS"])
def api_sync_verified_qty():
    if request.method == "OPTIONS":
        return "", 204

    token = request.headers.get("X-Sync-Token", "")
    if not token or not secrets.compare_digest(token, _load_or_create_sync_token()):
        return jsonify({"error": "同步碼不對，請確認腳本設定的 token。"}), 401

    payload = request.get_json(silent=True) or {}
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        return jsonify({"error": "沒有帶任何品項資料。"}), 400
    operator = norm_text(payload.get("operator")) or "酷澎驗收同步"

    stamp = now()
    matched = 0
    not_found = []

    conn = get_conn()
    try:
        for item in items:
            po_number = norm_key(item.get("po_number"))
            sku_id = norm_key(item.get("sku_id"))
            # 實際驗入數量＝酷澎後台「收貨數量」(receivedQty)。舊版腳本用
            # confirmed_qty 這個鍵送，保留相容讀取，避免快取到舊腳本時直接壞掉。
            raw_qty = item.get("verified_qty")
            if raw_qty is None:
                raw_qty = item.get("confirmed_qty")
            qty = norm_int(raw_qty)
            if not po_number or not sku_id or qty is None:
                continue

            order = conn.execute(
                "SELECT * FROM order_rows WHERE po_number = ? AND sku_id = ?",
                (po_number, sku_id)).fetchone()
            if order is None:
                not_found.append(f"{po_number} / {sku_id}")
                continue
            order = dict(order)
            matched += 1

            if not _same(order["actual_verified_qty"], qty):
                conn.execute(
                    """UPDATE orders SET actual_verified_qty = ?, actual_verified_at = ?,
                           updated_at = ?, version = version + 1
                       WHERE id = ?""",
                    (qty, stamp, stamp, order["id"]))
                log_change(conn, order, "actual_verified_qty", "實際驗入數量",
                           order["actual_verified_qty"], qty, operator, "system",
                           "酷澎後台驗收工具同步")

            # 出貨數量 − 實際驗入數量 > 0，代表少到貨，自動把「短驗 差額」
            # 補進驗收註記——已經有內容就接在後面補一行，不覆蓋原本寫的；
            # 同一句話已經出現過就不再重複補，不然每次跑工具都會多一行。
            shortfall = (order["qty_ship"] or 0) - qty
            if shortfall > 0:
                note_text = f"短驗 {shortfall}"
                old_note = order["receiving_note"] or ""
                if note_text not in old_note:
                    new_note = f"{old_note}；{note_text}" if old_note else note_text
                    conn.execute(
                        "UPDATE orders SET receiving_note = ?, updated_at = ? WHERE id = ?",
                        (new_note, stamp, order["id"]))
                    log_change(conn, order, "receiving_note", "驗收註記",
                               old_note, new_note, operator, "system",
                               "系統依實際驗入數量自動判斷")
        conn.commit()
    finally:
        conn.close()

    return jsonify({
        "ok": True, "matched": matched, "not_found": not_found,
        "message": f"同步完成，比對到 {matched} 個品項"
                   + (f"，{len(not_found)} 個系統裡還沒有、略過" if not_found else ""),
    })


# ---------------------------------------------------------------- 歷程

def log_change(conn, order, field, label, old, new, operator, source, note=""):
    conn.execute(
        """INSERT INTO edit_logs
           (order_id, po_number, sku_id, field, field_label, old_value,
            new_value, operator, source, note, changed_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (order["id"], order["po_number"], order["sku_id"], field, label,
         "" if old is None else str(old), "" if new is None else str(new),
         operator, source, note, now()),
    )


# ---------------------------------------------------------------- 頁面

@app.route("/")
def index():
    return render_template(
        "index.html", build_version=BUILD_VERSION,
        logged_in_user=session.get("user", ""),
    )


@app.route("/api/config")
def api_config():
    cfg = get_config()
    profiles = get_profiles()
    conn = get_conn()
    try:
        def distinct(column):
            rows = conn.execute(
                f"SELECT DISTINCT {column} AS v FROM order_rows "
                f"WHERE {column} IS NOT NULL AND {column} != '' ORDER BY v"
            ).fetchall()
            return [r["v"] for r in rows]

        # 這裡不再回傳 operators：操作人員已經改成「登入的是誰就是誰」，
        # config.json 裡那份名單既不影響登入、也不影響歷程，留著只會讓人
        # 以為改了就有用（之前就發生過「為什麼名單裡看不到某人」的誤會）。
        return jsonify({
            "po_statuses": cfg.get("po_statuses", []),
            "receiving_statuses": cfg.get("receiving_statuses", []),
            "order_types": cfg.get("order_types", []),
            "shipping_methods": cfg.get("shipping_methods", []),
            "brands": distinct("brand"),
            "lines": distinct("line"),
            "warehouses": distinct("warehouse"),
            "export_profiles": [
                {"key": k, "label": v.get("label", k), "note": v.get("note", ""),
                 "mark_pulled_default": bool(v.get("mark_pulled_default", False)),
                 "column_count": len(v.get("columns", []))}
                for k, v in profiles.items()
            ],
        })
    finally:
        conn.close()


# ---------------------------------------------------------------- 查詢

def build_filter(args):
    where, params = [], []

    def add(clause, *values):
        where.append(clause)
        params.extend(values)

    po = norm_text(args.get("po_number"))
    if po:
        add("po_number LIKE ?", f"%{po}%")

    keyword = norm_text(args.get("keyword"))
    if keyword:
        add("(product_name LIKE ? OR yf_sku LIKE ? OR sku_id LIKE ? OR barcode LIKE ?)",
            *[f"%{keyword}%"] * 4)

    # 篩選選單把「CPG-潔品」「CPG-紙品」這類 CPG 開頭的線別合併成一個
    # 「CPG」選項，選了它要撈出所有 CPG- 開頭的線別，不是精準比對。
    line = norm_text(args.get("line"))
    if line == "CPG":
        add("line LIKE ?", "CPG-%")
    elif line:
        add("line = ?", line)

    for column, key in (("brand", "brand"), ("warehouse", "warehouse"),
                        ("po_status", "po_status"),
                        ("receiving_status", "receiving_status"),
                        ("order_type", "order_type"),
                        ("shipping_method", "shipping_method")):
        value = norm_text(args.get(key))
        if value:
            add(f"{column} = ?", value)

    if args.get("flagged") == "1":
        add("flagged = 1")

    date_from = norm_date(args.get("date_from"))
    if date_from:
        add("delivery_date >= ?", date_from)
    date_to = norm_date(args.get("date_to"))
    if date_to:
        add("delivery_date <= ?", date_to)

    pulled = args.get("is_pulled")
    if pulled in ("0", "1"):
        add("is_pulled = ?", int(pulled))

    if args.get("needs_review") == "1":
        add("needs_review = 1")

    if args.get("alert_level"):
        add("alert_level = ?", norm_text(args.get("alert_level")))

    clause = (" WHERE " + " AND ".join(where)) if where else ""
    return clause, params


@app.route("/api/orders")
def api_orders():
    clause, params = build_filter(request.args)
    page = max(1, norm_int(request.args.get("page")) or 1)
    size = min(500, max(10, norm_int(request.args.get("page_size")) or 100))

    conn = get_conn()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) AS c FROM order_rows{clause}", params
        ).fetchone()["c"]

        summary = conn.execute(
            f"""SELECT COALESCE(SUM(qty_ship),0) AS qty,
                       SUM(CASE WHEN needs_review=1 THEN 1 ELSE 0 END) AS review,
                       SUM(CASE WHEN alert_level='changed_after_pull' THEN 1 ELSE 0 END) AS after_pull,
                       SUM(CASE WHEN is_pulled=1 THEN 1 ELSE 0 END) AS pulled
                FROM order_rows{clause}""",
            params,
        ).fetchone()

        rows = conn.execute(
            f"""SELECT * FROM order_rows{clause}
                ORDER BY needs_review DESC,
                         CASE alert_level WHEN 'changed_after_pull' THEN 0
                                          WHEN 'changed' THEN 1 ELSE 2 END,
                         delivery_date, po_number, seq_no, sku_id
                LIMIT ? OFFSET ?""",
            params + [size, (page - 1) * size],
        ).fetchall()

        return jsonify({
            "total": total,
            "page": page,
            "page_size": size,
            "summary": {
                "qty_ship": summary["qty"] or 0,
                "needs_review": summary["review"] or 0,
                "changed_after_pull": summary["after_pull"] or 0,
                "pulled": summary["pulled"] or 0,
            },
            "rows": [dict(r) for r in rows],
        })
    finally:
        conn.close()


@app.route("/api/orders/<int:order_id>/logs")
def api_order_logs(order_id):
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM edit_logs WHERE order_id = ? ORDER BY changed_at DESC, id DESC",
            (order_id,),
        ).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


def build_log_filter(args):
    """歷程查詢條件。日期用 changed_at 前綴比對，避免時分秒干擾。"""
    where, params = [], []

    def add(clause, *values):
        where.append(clause)
        params.extend(values)

    if norm_text(args.get("po_number")):
        add("po_number LIKE ?", f"%{norm_text(args.get('po_number'))}%")
    if norm_text(args.get("field")):
        add("field = ?", norm_text(args.get("field")))
    if norm_text(args.get("operator")):
        add("operator = ?", norm_text(args.get("operator")))
    if norm_text(args.get("source")):
        add("source = ?", norm_text(args.get("source")))
    if norm_text(args.get("keyword")):
        kw = f"%{norm_text(args.get('keyword'))}%"
        add("(old_value LIKE ? OR new_value LIKE ? OR note LIKE ? OR sku_id LIKE ?)",
            kw, kw, kw, kw)
    if norm_date(args.get("date_from")):
        add("changed_at >= ?", norm_date(args.get("date_from")) + " 00:00:00")
    if norm_date(args.get("date_to")):
        add("changed_at <= ?", norm_date(args.get("date_to")) + " 23:59:59")

    return ((" WHERE " + " AND ".join(where)) if where else ""), params


@app.route("/api/logs")
def api_logs():
    """全域歷程檢視：誰在什麼時候改了什麼，可依單號／欄位／人員／期間篩。"""
    clause, params = build_log_filter(request.args)
    page = max(1, norm_int(request.args.get("page")) or 1)
    size = min(500, max(10, norm_int(request.args.get("page_size")) or 100))

    conn = get_conn()
    try:
        total = conn.execute(
            f"SELECT COUNT(*) AS c FROM edit_logs{clause}", params).fetchone()["c"]
        rows = conn.execute(
            f"""SELECT * FROM edit_logs{clause}
                ORDER BY changed_at DESC, id DESC LIMIT ? OFFSET ?""",
            params + [size, (page - 1) * size]).fetchall()

        # 篩選面板用的選項，只列真的出現過的值
        fields = conn.execute(
            "SELECT DISTINCT field, field_label FROM edit_logs ORDER BY field_label"
        ).fetchall()
        operators = conn.execute(
            "SELECT DISTINCT operator FROM edit_logs WHERE operator != '' ORDER BY operator"
        ).fetchall()

        return jsonify({
            "total": total, "page": page, "page_size": size,
            "rows": [dict(r) for r in rows],
            "fields": [{"field": f["field"], "label": f["field_label"]} for f in fields],
            "operators": [o["operator"] for o in operators],
        })
    finally:
        conn.close()


@app.route("/api/logs/export", methods=["POST"])
def api_logs_export():
    """把目前篩選出來的歷程匯成 Excel。

    這張表最終是拿去跟酷澎對帳、釐清倉庫出錯責任用的，所以要能整份帶走，
    不能只留在畫面上一筆一筆看。
    """
    payload = request.get_json(silent=True) or {}
    operator = current_operator()

    clause, params = build_log_filter(payload.get("filters") or {})
    conn = get_conn()
    try:
        rows = conn.execute(
            f"SELECT * FROM edit_logs{clause} ORDER BY changed_at DESC, id DESC "
            f"LIMIT 50000", params).fetchall()
    finally:
        conn.close()

    if not rows:
        return jsonify({"error": "目前的篩選條件沒有任何歷程可以匯出。"}), 400

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "修改歷程"
    headers = ["時間", "PO 單號", "SKU ID", "層級", "欄位", "改前", "改後",
               "操作人員", "來源", "備註"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="161D29")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    source_zh = {"import": "匯入", "manual": "手動", "system": "系統"}
    for r in rows:
        ws.append([
            r["changed_at"], r["po_number"], r["sku_id"],
            "整張單" if not r["sku_id"] else "單一品項",
            r["field_label"], r["old_value"], r["new_value"],
            r["operator"], source_zh.get(r["source"], r["source"]), r["note"],
        ])

    for idx, width in enumerate([19, 18, 18, 10, 12, 22, 22, 10, 8, 30], start=1):
        letter = ws.cell(row=1, column=idx).column_letter
        ws.column_dimensions[letter].width = width
        if idx in (2, 3):                      # 長數字一律當文字，避免科學記號
            for cell in ws[letter][1:]:
                cell.number_format = "@"
    ws.freeze_panes = "A2"

    stamp = db.file_stamp()
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream, as_attachment=True, download_name=f"修改歷程_{stamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------- 線上編輯

@app.route("/api/orders/<int:order_id>", methods=["PUT"])
def api_update_order(order_id):
    payload = request.get_json(silent=True) or {}
    operator = current_operator()

    client_version = norm_int(payload.get("version"))
    if client_version is None:
        return jsonify({"error": "缺少版本資訊，請重新整理後再試。"}), 400

    conn = get_conn()
    try:
        order = conn.execute("SELECT * FROM order_rows WHERE id = ?", (order_id,)).fetchone()
        if order is None:
            return jsonify({"error": "找不到這筆訂單。"}), 404
        order = dict(order)

        # 樂觀鎖：三個人共用一份資料，last-write-wins 會無聲蓋掉同事的修改
        if order["version"] != client_version:
            return jsonify({
                "error": "conflict",
                "message": (
                    f"這筆訂單剛剛被「{order['updated_at']}」的另一次儲存修改過，"
                    "你看到的已經不是最新版本。請重新載入後再編輯。"
                ),
                "current": order,
            }), 409

        # 拉單鎖定只擋「會影響倉庫出貨的數字」（目前只有出貨數量）；
        # 備註／驗收註記是事後才填的資訊，永遠不受這個鎖影響，所以只看
        # 這次要改的欄位裡，扣掉那兩個豁免欄位後還有沒有剩下的。
        locked_fields = (set(payload) & set(EDITABLE_FIELDS)) - ALWAYS_EDITABLE_FIELDS
        if order["is_pulled"] and locked_fields and not payload.get("force_edit"):
            return jsonify({
                "error": "locked",
                "message": "這筆訂單已拉單並鎖定。若確實需要修改，請先解除拉單鎖定。",
            }), 423

        changes = []
        for field, label in EDITABLE_FIELDS.items():
            if field not in payload:
                continue
            raw = payload[field]
            new_val = norm_int(raw) if field == "qty_ship" else norm_text(raw)

            old_val = order[field]
            if _same(old_val, new_val):
                continue
            changes.append((field, label, old_val, new_val))

        # 補空白：只有原本沒有值的欄位才收，已經有值的直接忽略
        for field, label in FILLABLE_FIELDS.items():
            if field not in payload:
                continue
            new_val = norm_text(payload[field])
            if not new_val:
                continue
            if norm_text(order[field]):
                return jsonify({
                    "error": "readonly",
                    "message": (
                        f"「{label}」已經有值（{order[field]}），不開放修改。"
                        "這個欄位以酷澎整合表為準，需要更正請回頭修主檔後重新上傳。"
                    ),
                }), 400
            changes.append((field, label, order[field], new_val))

        if not changes:
            return jsonify({"ok": True, "changed": 0, "order": order})

        sets = ", ".join(f"{f} = ?" for f, _, _, _ in changes)
        values = [v for _, _, _, v in changes]
        cur = conn.execute(
            f"""UPDATE orders SET {sets}, updated_at = ?, version = version + 1
                WHERE id = ? AND version = ?""",
            values + [now(), order_id, client_version],
        )
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"error": "conflict",
                            "message": "儲存瞬間有其他人也改了這筆，請重新載入。"}), 409

        for field, label, old_val, new_val in changes:
            log_change(conn, order, field, label, old_val, new_val, operator, "manual")
            if field in SKU_OVERRIDE_FIELDS:
                conn.execute(f"UPDATE orders SET {field}_overridden = 1 WHERE id = ?",
                             (order_id,))
        conn.commit()

        updated = dict(conn.execute("SELECT * FROM order_rows WHERE id = ?",
                                    (order_id,)).fetchone())
        return jsonify({"ok": True, "changed": len(changes), "order": updated})
    finally:
        conn.close()


# ---------------------------------------------------------------- PO 層級

def log_po_change(conn, po_number, field, label, old, new, operator,
                  source="manual", note=""):
    """整張 PO 的變更只記一筆，sku_id 留空代表『這是整張單的事』。"""
    conn.execute(
        """INSERT INTO edit_logs
           (order_id, po_number, sku_id, field, field_label, old_value,
            new_value, operator, source, note, changed_at)
           VALUES (0,?,'',?,?,?,?,?,?,?,?)""",
        (po_number, field, label,
         "" if old is None else str(old), "" if new is None else str(new),
         operator, source, note, now()),
    )


def po_summary_sql(clause):
    """把 SKU 收攏成一列一張 PO。

    交期／倉別在同一張 PO 底下本來就一致（實際資料驗證過），用 MIN 取值
    即可；品牌則可能混多個，用逗號串起來，這樣首頁篩品牌時混單也篩得到。

    篩選條件只用來決定「哪幾張 PO 入選」，統計值一律涵蓋該 PO 底下的
    全部品項。否則搜「whiskas」時，那張含 5 個品牌的單會只顯示 whiskas、
    品項數與數量也只算到符合的那幾項，資訊是錯的。
    """
    # GROUP_CONCAT 是 SQLite 的寫法，PostgreSQL 要用 STRING_AGG，
    # 兩邊語法不同、但效果一樣（DISTINCT 值用逗號串起來）。
    concat = (lambda col: f"STRING_AGG(DISTINCT {col}, ',')") if db.IS_POSTGRES \
        else (lambda col: f"GROUP_CONCAT(DISTINCT {col})")
    return f"""
        SELECT po_number,
               MIN(order_type)      AS order_type,
               MIN(parent_po)       AS parent_po,
               MIN(po_status)       AS po_status,
               MIN(receiving_status) AS receiving_status,
               MAX(is_pulled)       AS is_pulled,
               MIN(pulled_at)       AS pulled_at,
               MIN(pulled_by)       AS pulled_by,
               MIN(filed_date)      AS filed_date,
               MIN(shipping_method) AS shipping_method,
               MAX(flagged)         AS flagged,
               MIN(delivery_date)   AS delivery_date,
               MIN(warehouse)       AS warehouse,
               {concat("line")}  AS lines_csv,
               {concat("brand")} AS brands_csv,
               COUNT(*)             AS sku_count,
               COALESCE(SUM(qty_coupang),0) AS qty_coupang,
               COALESCE(SUM(qty_ship),0)    AS qty_ship,
               -- 這裡故意不 COALESCE 成 0：一個品項都還沒同步過的單，
               -- 總和該是「還沒有數字」（畫面上顯示 —），不是「驗入 0 件」，
               -- 兩者意思差很多，混在一起會誤導人。
               SUM(actual_verified_qty)     AS actual_verified_qty,
               SUM(CASE WHEN needs_review=1 THEN 1 ELSE 0 END) AS review_count,
               SUM(CASE WHEN alert_level='changed_after_pull' THEN 1 ELSE 0 END)
                                            AS after_pull_count,
               MIN(po_version)      AS po_version
        FROM order_rows
        WHERE po_number IN (SELECT po_number FROM order_rows{clause})
        GROUP BY po_number
    """


def _csv_clean(value):
    """GROUP_CONCAT 出來的字串去掉空值、排序後用逗號串好。"""
    if not value:
        return ""
    parts = sorted({p.strip() for p in value.split(",") if p.strip()})
    return ", ".join(parts)


@app.route("/api/pos")
def api_pos():
    """首頁：一列一張 PO。"""
    clause, params = build_filter(request.args)
    page = max(1, norm_int(request.args.get("page")) or 1)
    size = min(500, max(10, norm_int(request.args.get("page_size")) or 50))

    conn = get_conn()
    try:
        base = po_summary_sql(clause)
        total = conn.execute(
            f"SELECT COUNT(*) AS c FROM ({base})", params).fetchone()["c"]

        summary = conn.execute(
            f"""SELECT COALESCE(SUM(qty_ship),0) AS qty,
                       SUM(CASE WHEN review_count>0 THEN 1 ELSE 0 END) AS review,
                       SUM(CASE WHEN after_pull_count>0 THEN 1 ELSE 0 END) AS after_pull,
                       SUM(CASE WHEN is_pulled=1 THEN 1 ELSE 0 END) AS pulled
                FROM ({base})""", params).fetchone()

        rows = conn.execute(
            f"""SELECT * FROM ({base})
                ORDER BY after_pull_count DESC, review_count DESC,
                         delivery_date, po_number
                LIMIT ? OFFSET ?""",
            params + [size, (page - 1) * size]).fetchall()

        out = []
        for r in rows:
            item = dict(r)
            item["brands"] = _csv_clean(item.pop("brands_csv"))
            item["lines"] = _csv_clean(item.pop("lines_csv"))
            out.append(item)

        return jsonify({
            "total": total, "page": page, "page_size": size,
            "summary": {
                "qty_ship": summary["qty"] or 0,
                "needs_review": summary["review"] or 0,
                "changed_after_pull": summary["after_pull"] or 0,
                "pulled": summary["pulled"] or 0,
            },
            "rows": out,
        })
    finally:
        conn.close()


@app.route("/api/pos/<po_number>")
def api_po_detail(po_number):
    """點開一張 PO：表頭 + 這張單所有 SKU + 整張單的歷程。"""
    po_number = norm_key(po_number)
    conn = get_conn()
    try:
        header = conn.execute(
            "SELECT * FROM po_headers WHERE po_number = ?", (po_number,)).fetchone()
        if header is None:
            return jsonify({"error": "找不到這張 PO。"}), 404

        skus = conn.execute(
            """SELECT * FROM order_rows WHERE po_number = ?
               ORDER BY seq_no, sku_id""", (po_number,)).fetchall()
        logs = conn.execute(
            """SELECT * FROM edit_logs WHERE po_number = ?
               ORDER BY changed_at DESC, id DESC LIMIT 300""", (po_number,)).fetchall()

        skus = [dict(s) for s in skus]
        return jsonify({
            "header": dict(header),
            "brands": _csv_clean(",".join(s["brand"] or "" for s in skus)),
            "lines": _csv_clean(",".join(s["line"] or "" for s in skus)),
            "delivery_date": skus[0]["delivery_date"] if skus else "",
            "warehouse": skus[0]["warehouse"] if skus else "",
            "skus": skus,
            "logs": [dict(l) for l in logs],
        })
    finally:
        conn.close()


@app.route("/api/pos/<po_number>", methods=["PUT"])
def api_update_po(po_number):
    """改整張 PO：狀態類欄位寫 po_headers，交期／倉別套用到全部 SKU。"""
    po_number = norm_key(po_number)
    payload = request.get_json(silent=True) or {}
    operator = current_operator()

    client_version = norm_int(payload.get("po_version"))
    if client_version is None:
        return jsonify({"error": "缺少版本資訊，請重新整理後再試。"}), 400

    conn = get_conn()
    try:
        header = conn.execute(
            "SELECT * FROM po_headers WHERE po_number = ?", (po_number,)).fetchone()
        if header is None:
            return jsonify({"error": "找不到這張 PO。"}), 404
        header = dict(header)

        if header["version"] != client_version:
            return jsonify({
                "error": "conflict",
                "message": (f"這張單剛剛被另一次儲存修改過（{header['updated_at']}），"
                            "你看到的不是最新版本，請重新載入後再編輯。"),
            }), 409

        stamp = now()
        head_changes = []
        for field, label in PO_EDITABLE_FIELDS.items():
            if field not in payload:
                continue
            new_val = (norm_date(payload[field]) if field == "filed_date"
                       else norm_text(payload[field]))
            if _same(header[field], new_val):
                continue
            head_changes.append((field, label, header[field], new_val))

        # 交期／倉別存在每一列 SKU 上（要參與匯入比對），但 OP 是整張單一起改
        row_changes = []
        first = conn.execute(
            "SELECT delivery_date, warehouse, order_type FROM orders "
            "WHERE po_number = ? LIMIT 1", (po_number,)).fetchone()
        if first is not None:
            for field, label in PO_COUPANG_FIELDS.items():
                if field not in payload:
                    continue
                new_val = _PO_COUPANG_FIELD_NORM[field](payload[field])
                if _same(first[field], new_val):
                    continue
                row_changes.append((field, label, first[field], new_val))

        if not head_changes and not row_changes:
            return jsonify({"ok": True, "changed": 0})

        # 鎖定檢查放在算完「真的有沒有異動」之後——這些欄位在畫面上鎖單
        # 時本來就是 disabled，正常不會送出真的改變；只有真的想繞過鎖定
        # 硬改時才擋下來。放在最前面會連「這張單其實什麼都沒改、只是
        # 品項明細裡動了驗收註記」這種情況也一起誤擋，导致存檔整包失敗。
        locked_fields = ({f for f, _, _, _ in head_changes} |
                         {f for f, _, _, _ in row_changes}) - PO_ALWAYS_EDITABLE_FIELDS
        if header["is_pulled"] and locked_fields and not payload.get("force_edit"):
            return jsonify({
                "error": "locked",
                "message": "這張單已拉單並鎖定。若確實需要修改，請先解除拉單鎖定。",
            }), 423

        if head_changes:
            sets = ", ".join(f"{f} = ?" for f, _, _, _ in head_changes)
            cur = conn.execute(
                f"""UPDATE po_headers SET {sets}, updated_at = ?, version = version + 1
                    WHERE po_number = ? AND version = ?""",
                [v for _, _, _, v in head_changes] + [stamp, po_number, client_version])
            if cur.rowcount == 0:
                conn.rollback()
                return jsonify({"error": "conflict",
                                "message": "儲存瞬間有其他人也改了這張單，請重新載入。"}), 409

        for field, label, old_val, new_val in row_changes:
            # 一併標記「這個欄位人工調整過」，之後匯入就不再覆蓋它
            conn.execute(
                f"""UPDATE orders SET {field} = ?, {field}_overridden = 1,
                        updated_at = ?, version = version + 1
                    WHERE po_number = ?""", (new_val, stamp, po_number))

        for field, label, old_val, new_val in head_changes + row_changes:
            log_po_change(conn, po_number, field, label, old_val, new_val, operator)
        conn.commit()
        return jsonify({"ok": True, "changed": len(head_changes) + len(row_changes)})
    finally:
        conn.close()


@app.route("/api/pos/review", methods=["POST"])
def api_clear_review():
    """OP 確認過異動後把整張單的警示燈關掉，每張單各記一筆歷程。"""
    payload = request.get_json(silent=True) or {}
    operator = current_operator()
    pos = [norm_key(p) for p in payload.get("po_numbers", []) if norm_key(p)]
    if not pos:
        return jsonify({"error": "沒有選取任何訂單。"}), 400

    conn = get_conn()
    try:
        placeholders = ",".join("?" * len(pos))
        rows = conn.execute(
            f"""SELECT po_number, COUNT(*) AS n FROM orders
                WHERE po_number IN ({placeholders}) AND needs_review = 1
                GROUP BY po_number""", pos).fetchall()
        for row in rows:
            log_po_change(conn, row["po_number"], "needs_review", "異動確認",
                          f"{row['n']} 項待確認", "已確認", operator)
        conn.execute(
            f"""UPDATE orders SET needs_review = 0, alert_level = '',
                   review_reason = '', updated_at = ?, version = version + 1
                WHERE po_number IN ({placeholders}) AND needs_review = 1""",
            [now()] + pos)
        conn.commit()
        return jsonify({"ok": True, "count": len(rows)})
    finally:
        conn.close()


@app.route("/api/pos/pull", methods=["POST"])
def api_set_pulled():
    """手動調整整張單的拉單狀態。

    標記／解除都不強制填理由——OP 反映這一步是例行操作，強制填理由
    反而拖慢流程。理由欄位保留、有填就照樣寫進歷程，只是不再擋下
    沒填理由的操作；誰在什麼時候改了拉單狀態，還是查得到。
    """
    payload = request.get_json(silent=True) or {}
    operator = current_operator()
    reason = norm_text(payload.get("reason"))
    pos = [norm_key(p) for p in payload.get("po_numbers", []) if norm_key(p)]
    target = 1 if payload.get("pulled") else 0
    if not pos:
        return jsonify({"error": "沒有選取任何訂單。"}), 400

    conn = get_conn()
    try:
        placeholders = ",".join("?" * len(pos))
        rows = conn.execute(
            f"""SELECT po_number, is_pulled FROM po_headers
                WHERE po_number IN ({placeholders}) AND is_pulled != ?""",
            pos + [target]).fetchall()
        stamp = now()
        for row in rows:
            log_po_change(conn, row["po_number"], "is_pulled", "拉單狀態",
                          "已拉單" if row["is_pulled"] else "未拉單",
                          "已拉單" if target else "未拉單",
                          operator, "manual", reason)
        if target:
            conn.execute(
                f"""UPDATE po_headers SET is_pulled = 1, pulled_at = ?, pulled_by = ?,
                       updated_at = ?, version = version + 1
                    WHERE po_number IN ({placeholders}) AND is_pulled = 0""",
                [stamp, operator, stamp] + pos)
        else:
            conn.execute(
                f"""UPDATE po_headers SET is_pulled = 0, pulled_at = '', pulled_by = '',
                       pulled_batch_id = NULL, updated_at = ?, version = version + 1
                    WHERE po_number IN ({placeholders}) AND is_pulled = 1""",
                [stamp] + pos)
        conn.commit()
        return jsonify({"ok": True, "count": len(rows)})
    finally:
        conn.close()


@app.route("/api/pos/flag", methods=["POST"])
def api_set_flagged():
    """切換個人標記（全公司共用一份、只有開／關兩種狀態）。

    純粹是 OP 自己想特別留意哪張單用的顏色標記，跟出貨、驗收都無關，
    所以不受拉單鎖定限制，也不記進 edit_logs 歷程（不是業務異動）。
    """
    payload = request.get_json(silent=True) or {}
    pos = [norm_key(p) for p in payload.get("po_numbers", []) if norm_key(p)]
    target = 1 if payload.get("flagged") else 0
    if not pos:
        return jsonify({"error": "沒有選取任何訂單。"}), 400

    conn = get_conn()
    try:
        placeholders = ",".join("?" * len(pos))
        cur = conn.execute(
            f"""UPDATE po_headers SET flagged = ?, updated_at = ?, version = version + 1
                WHERE po_number IN ({placeholders}) AND flagged != ?""",
            [target, now()] + pos + [target])
        conn.commit()
        return jsonify({"ok": True, "count": cur.rowcount})
    finally:
        conn.close()


@app.route("/api/pos/status", methods=["POST"])
def api_batch_status():
    """批次改整批 PO 的狀態，每張單各記一筆歷程（不是只記一筆批次）。"""
    payload = request.get_json(silent=True) or {}
    operator = current_operator()
    pos = [norm_key(p) for p in payload.get("po_numbers", []) if norm_key(p)]
    field = norm_text(payload.get("field"))
    value = norm_text(payload.get("value"))
    if not pos:
        return jsonify({"error": "沒有選取任何訂單。"}), 400
    if field not in ("po_status", "receiving_status", "shipping_method"):
        return jsonify({"error": "只能批次修改 PO 狀態、驗收狀態或配送方式。"}), 400
    if not value:
        return jsonify({"error": "請選擇要改成什麼狀態。"}), 400

    label = PO_EDITABLE_FIELDS[field]
    conn = get_conn()
    try:
        placeholders = ",".join("?" * len(pos))
        rows = conn.execute(
            f"""SELECT po_number, {field} AS old FROM po_headers
                WHERE po_number IN ({placeholders}) AND {field} != ?""",
            pos + [value]).fetchall()
        for row in rows:
            log_po_change(conn, row["po_number"], field, label,
                          row["old"], value, operator)
        conn.execute(
            f"""UPDATE po_headers SET {field} = ?, updated_at = ?, version = version + 1
                WHERE po_number IN ({placeholders})""",
            [value, now()] + pos)
        conn.commit()
        return jsonify({"ok": True, "count": len(rows)})
    finally:
        conn.close()


# ---------------------------------------------------------------- 匯入

@app.route("/api/import/preview", methods=["POST"])
def api_import_preview():
    """第一階段：只解析與比對，不寫入 orders。OP 看完報告才決定。"""
    operator = current_operator()
    upload = request.files.get("file")
    if upload is None or not upload.filename:
        return jsonify({"error": "沒有收到檔案。"}), 400

    try:
        data = io.BytesIO(upload.read())
        rows, warnings = parse_workbook(data, upload.filename)
    except ImportError_ as exc:
        return jsonify({"error": str(exc)}), 400

    conn = get_conn()
    try:
        result = diff_rows(conn, rows)
        preview = {
            "filename": upload.filename,
            "operator": operator,
            "warnings": warnings,
            "new": result["new"],
            "updated": result["updated"],
            "removed": result["removed"],
            "identical_count": len(result["identical"]),
            "identical_keys": [[r["po_number"], r["sku_id"]]
                               for r in result["identical"]],
        }
        cur = conn.execute(
            """INSERT INTO import_batches
               (filename, operator, rows_total, rows_new, rows_updated,
                rows_identical, rows_error, committed, preview_json, created_at)
               VALUES (?,?,?,?,?,?,?,0,?,?)""",
            (upload.filename, operator, len(rows), len(result["new"]),
             len(result["updated"]), len(result["identical"]), len(warnings),
             json.dumps(preview, ensure_ascii=False), now()),
        )
        conn.commit()
        batch_id = cur.lastrowid
    finally:
        conn.close()

    after_pull = [u for u in result["updated"] if u["after_pull"]]
    return jsonify({
        "batch_id": batch_id,
        "filename": upload.filename,
        "rows_total": len(rows),
        "new_count": len(result["new"]),
        "updated_count": len(result["updated"]),
        "identical_count": len(result["identical"]),
        "after_pull_count": len(after_pull),
        "removed_count": len(result["removed"]),
        "warnings": warnings,
        "new_preview": result["new"][:200],
        "updated": result["updated"][:200],
        "removed": result["removed"][:200],
    })


@app.route("/api/import/commit", methods=["POST"])
def api_import_commit():
    """第二階段：全有全無寫入。中間任何一列出錯就整批 rollback。"""
    payload = request.get_json(silent=True) or {}
    batch_id = norm_int(payload.get("batch_id"))
    operator = current_operator()
    if not batch_id:
        return jsonify({"error": "缺少批次資訊，請重新上傳。"}), 400

    conn = get_conn()
    try:
        batch = conn.execute("SELECT * FROM import_batches WHERE id = ?",
                             (batch_id,)).fetchone()
        if batch is None:
            return jsonify({"error": "找不到這個匯入批次，請重新上傳。"}), 404
        if batch["committed"]:
            return jsonify({"error": "這個批次已經匯入過了，請重新上傳檔案。"}), 400

        preview = json.loads(batch["preview_json"])
    finally:
        conn.close()

    db.backup_db("import")   # 寫入前先留一份身家

    conn = get_conn()
    try:
        # BEGIN IMMEDIATE 是 SQLite 專屬語法，搶先拿寫入鎖避免之後升級
        # 鎖時撞死結；PostgreSQL 沒有這個問題（MVCC），且每個敘述本來
        # 就已經在交易裡，不需要也不能下這行。
        if not db.IS_POSTGRES:
            conn.execute("BEGIN IMMEDIATE")
        stamp = now()
        inserted = updated = 0

        today = db.today()
        # INSERT OR IGNORE 是 SQLite 寫法，PostgreSQL 要用 ON CONFLICT
        # DO NOTHING，效果一樣：這張 PO 表頭已經存在就什麼都不做。
        insert_po_header = (
            """INSERT INTO po_headers
               (po_number, po_status, receiving_status, is_pulled,
                filed_date, created_at, updated_at)
               VALUES (?, '已建立', '未驗收', 0, ?, ?, ?)
               ON CONFLICT (po_number) DO NOTHING"""
            if db.IS_POSTGRES else
            """INSERT OR IGNORE INTO po_headers
               (po_number, po_status, receiving_status, is_pulled,
                filed_date, created_at, updated_at)
               VALUES (?, '已建立', '未驗收', 0, ?, ?, ?)"""
        )
        for row in preview["new"]:
            # PO 表頭只在第一次見到這張單時建立。之後同一張單再上傳，
            # 這裡什麼都不做——狀態與建檔日一律以系統為準，不被 Excel 覆蓋。
            conn.execute(insert_po_header, (row["po_number"], today, stamp, stamp))

            columns = ", ".join(INSERT_FIELDS)
            marks = ", ".join("?" * len(INSERT_FIELDS))
            values = [row.get(f) for f in INSERT_FIELDS]
            cur = conn.execute(
                f"""INSERT INTO orders ({columns}, qty_ship,
                        source_file, first_seen_at, last_seen_at,
                        created_at, updated_at)
                    VALUES ({marks}, ?, ?, ?, ?, ?, ?)""",
                values + [desired_ship(row),
                          preview["filename"], stamp, stamp, stamp, stamp],
            )
            order = {"id": cur.lastrowid, "po_number": row["po_number"],
                     "sku_id": row["sku_id"]}
            log_change(conn, order, "_created", "新增訂單", "",
                       f"{row['po_number']} / {row['sku_id']}", operator,
                       "import", preview["filename"])
            inserted += 1

        for item in preview["updated"]:
            row = item["row"]
            order_id = item["order_id"]
            current = conn.execute("SELECT * FROM order_rows WHERE id = ?",
                                   (order_id,)).fetchone()
            if current is None:
                continue
            current = dict(current)

            sets, values = [], []
            critical_hit = False
            unsynced = []
            for change in item["changes"]:
                field = change["field"]
                # OP 已經人工調整過的交期／倉別不覆蓋——那是跟酷澎談好的
                # 結果，酷澎後台只是還沒更新。但差異照樣記歷程、照樣亮燈，
                # 不會把「兩邊還沒同步」這件事藏起來。
                if current.get(f"{field}_overridden"):
                    unsynced.append(f"{change['label']}（酷澎仍為 {change['new']}）")
                    log_change(conn, current, field, change["label"],
                               change["new"], change["old"], operator, "import",
                               "酷澎檔案仍是舊值，保留人工調整結果不覆蓋")
                    critical_hit = critical_hit or field in CRITICAL_FIELDS
                    continue
                sets.append(f"{field} = ?")
                values.append(row.get(field))
                log_change(conn, current, field, change["label"],
                           change["old"], change["new"], operator, "import",
                           preview["filename"])
                if field in CRITICAL_FIELDS:
                    critical_hit = True

            # 出貨數量若 OP 從未手動改過，就跟著整合表走——整合表如果
            # 自己就帶了「出貨數量」欄，那才是真正要出的量，比死板地
            # 拿下單數量複製過去準；檔案沒帶這欄才退回用下單數量頂著
            # （desired_ship 已經處理這個順序）。一旦手動調整過，匯入
            # 永遠不再碰它，但差異照樣要記歷程、要亮燈，不能藏起來。
            if item.get("ship_changed"):
                ship_note = ("隨整合表出貨數量連動" if row.get("qty_file_ship") is not None
                             else "隨下單數量連動")
                new_ship = item["desired_ship"]
                if not current["qty_ship_overridden"]:
                    sets.append("qty_ship = ?")
                    values.append(new_ship)
                    log_change(conn, current, "qty_ship", "出貨數量",
                               current["qty_ship"], new_ship,
                               operator, "import", ship_note)
                else:
                    unsynced.append(f"出貨數量（整合表為 {new_ship}）")
                    log_change(conn, current, "qty_ship", "出貨數量",
                               new_ship, current["qty_ship"], operator, "import",
                               "整合表出貨數量已變，保留人工調整結果不覆蓋")
                critical_hit = True

            reason_parts = [
                f"{c['label']} {c['old']}→{c['new']}" for c in item["changes"]
                if not current.get(f"{c['field']}_overridden")
            ]
            # 出貨數量沒有現成的「舊值→新值」文字（它不是 item["changes"]
            # 裡的一員），只有真的同步了才需要另外補一句，不然一張單只有
            # 出貨數量變動時，警示燈亮了但理由欄卻是空的，看不出為什麼。
            if item.get("ship_changed") and not current["qty_ship_overridden"]:
                reason_parts.append(f"出貨數量 {current['qty_ship']}→{item['desired_ship']}")
            # 這個品項之前被判定「酷澎移除」，這次匯入又出現了——解除標記，
            # 讓它之後照正常的異動流程走，不再排除於 PO 總數之外。
            if item.get("revived"):
                sets.append("removed_from_coupang = 0")
                log_change(conn, current, "removed_from_coupang", "酷澎移除狀態",
                           "已從酷澎移除", "重新出現在整合表", operator, "import",
                           preview["filename"])
                reason_parts.append("這個品項先前被判定已從酷澎移除，這次匯入重新出現")
                critical_hit = True
            reason = "、".join(reason_parts)
            if unsynced:
                reason = ("【與酷澎尚未同步】" + "、".join(unsynced)
                          + ("；" + reason if reason else ""))
            if current["is_pulled"] and critical_hit:
                alert = "changed_after_pull"
                reason = "【已拉單後遭變更】" + reason
            else:
                alert = "changed"

            sets += ["needs_review = 1", "alert_level = ?", "review_reason = ?",
                     "last_seen_at = ?", "updated_at = ?", "source_file = ?",
                     "version = version + 1"]
            values += [alert, reason[:900], stamp, stamp, preview["filename"]]

            conn.execute(f"UPDATE orders SET {', '.join(sets)} WHERE id = ?",
                         values + [order_id])
            updated += 1

        removed_count = 0
        for rrow in preview.get("removed", []):
            order_id = rrow["id"]
            current = conn.execute("SELECT * FROM order_rows WHERE id = ?",
                                   (order_id,)).fetchone()
            # 上一批次可能已經處理過同一筆（同一個檔案裡本來就不該重複，
            # 但保險起見還是查一次現況，避免 UPDATE 到不存在或已經改過
            # 標記的列）。
            if current is None or current["removed_from_coupang"]:
                continue
            current = dict(current)

            log_change(conn, current, "qty_ship", "出貨數量",
                       current["qty_ship"], 0, operator, "import",
                       "酷澎後台已移除此品項（下單數量歸零後，整合表不會再帶這一列），"
                       "出貨數量歸零、不計入 PO 總數，保留成稽核紀錄")

            reason = ("這個品項已在酷澎後台被移除，本次上傳的整合表已找不到這筆資料，"
                      "出貨數量已歸零、不計入 PO 總數")
            if current["is_pulled"]:
                alert = "changed_after_pull"
                reason = "【已拉單後遭變更】" + reason
            else:
                alert = "missing"

            conn.execute(
                """UPDATE orders SET qty_ship = 0, qty_ship_overridden = 0,
                       removed_from_coupang = 1, needs_review = 1,
                       alert_level = ?, review_reason = ?,
                       last_seen_at = ?, updated_at = ?, source_file = ?,
                       version = version + 1
                   WHERE id = ?""",
                (alert, reason[:900], stamp, stamp, preview["filename"], order_id),
            )
            removed_count += 1

        # 內容完全相同的列只更新「最後出現時間」，不動任何業務欄位、
        # 不寫歷程、不改 version —— 這就是「靜默去重」。
        conn.executemany(
            "UPDATE orders SET last_seen_at = ? WHERE po_number = ? AND sku_id = ?",
            [(stamp, po, sku) for po, sku in preview.get("identical_keys", [])],
        )

        conn.execute(
            "UPDATE import_batches SET committed = 1, committed_at = ? WHERE id = ?",
            (stamp, batch_id),
        )
        conn.commit()
    except Exception as exc:  # noqa: BLE001
        conn.rollback()
        return jsonify({"error": f"匯入失敗，資料已全部還原，沒有寫入任何一筆：{exc}"}), 500
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "inserted": inserted,
        "updated": updated,
        "identical": preview["identical_count"],
        "removed": removed_count,
    })


# ---------------------------------------------------------------- 匯出

def _format_cell(value, fmt):
    if value is None:
        return ""
    if fmt == "int":
        return norm_int(value)
    if fmt == "decimal":
        try:
            return float(value)
        except (TypeError, ValueError):
            return ""
    if fmt == "date":
        return norm_date(value)
    if fmt == "yesno":
        return "是" if norm_int(value) else "否"
    return value


@app.route("/api/export", methods=["POST"])
def api_export():
    """依畫面上的篩選條件匯出。

    是否標記為「已拉單」由 OP 明確決定：試算用的匯出不該把單鎖掉。
    """
    payload = request.get_json(silent=True) or {}
    operator = current_operator()
    profile_key = norm_text(payload.get("profile")) or "warehouse"
    mark_pulled = bool(payload.get("mark_pulled"))
    filters = payload.get("filters") or {}
    selected_pos = [norm_key(x) for x in payload.get("po_numbers", []) if norm_key(x)]


    profiles = get_profiles()
    profile = profiles.get(profile_key)
    if not profile:
        return jsonify({"error": f"找不到匯出格式「{profile_key}」，請檢查 export_profiles.json。"}), 400

    conn = get_conn()
    try:
        if selected_pos:
            placeholders = ",".join("?" * len(selected_pos))
            rows = conn.execute(
                f"SELECT * FROM order_rows WHERE po_number IN ({placeholders}) "
                f"ORDER BY delivery_date, po_number, seq_no, sku_id",
                selected_pos,
            ).fetchall()
            filter_desc = {"po_numbers": selected_pos}
        else:
            clause, params = build_filter(filters)
            rows = conn.execute(
                f"SELECT * FROM order_rows{clause} "
                f"ORDER BY delivery_date, po_number, seq_no, sku_id", params
            ).fetchall()
            filter_desc = filters

        if not rows:
            return jsonify({"error": "目前的篩選條件沒有任何資料可以匯出。"}), 400

        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = profile.get("label", profile_key)[:28]

        columns = profile.get("columns", [])
        headers = [c.get("header", c.get("field", "")) for c in columns]
        ws.append(headers)
        head_fill = PatternFill("solid", fgColor="1F3A5F")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in rows:
            record = dict(row)
            line = []
            for col in columns:
                if "const" in col:
                    line.append(col["const"])
                    continue
                value = record.get(col.get("field"))
                line.append(_format_cell(value, col.get("format")))
            ws.append(line)

        # 長 PO 單號/料號一定要當文字，否則 Excel 會轉成科學記號
        for idx, col in enumerate(columns, start=1):
            letter = ws.cell(row=1, column=idx).column_letter
            if col.get("format") == "text":
                for cell in ws[letter][1:]:
                    cell.number_format = "@"
            widths = {"text": 18, "date": 12, "int": 10, "decimal": 12}
            ws.column_dimensions[letter].width = widths.get(col.get("format"), 16)
        ws.freeze_panes = "A2"

        stamp = db.file_stamp()
        filename = f"酷澎出貨表_{profile_key}_{stamp}.xlsx"

        cur = conn.execute(
            """INSERT INTO export_batches
               (operator, profile, filename, row_count, mark_pulled,
                filter_json, created_at)
               VALUES (?,?,?,?,?,?,?)""",
            (operator, profile_key, filename, len(rows), int(mark_pulled),
             json.dumps(filter_desc, ensure_ascii=False), now()),
        )
        batch_id = cur.lastrowid
        conn.executemany(
            """INSERT INTO export_batch_items
               (batch_id, order_id, po_number, sku_id, qty_ship)
               VALUES (?,?,?,?,?)""",
            [(batch_id, r["id"], r["po_number"], r["sku_id"], r["qty_ship"])
             for r in rows],
        )

        if mark_pulled:
            # 拉單是整張 PO 的事：匯出裡只要有這張單的任何一個品項，
            # 整張單就算交出去了。
            stamp_now = now()
            pos = sorted({r["po_number"] for r in rows if not r["is_pulled"]})
            note = f"匯出批次 #{batch_id}／{profile.get('label', profile_key)}"
            for po in pos:
                log_po_change(conn, po, "is_pulled", "拉單狀態", "未拉單",
                              "已拉單", operator, "system", note)
            if pos:
                placeholders = ",".join("?" * len(pos))
                conn.execute(
                    f"""UPDATE po_headers SET is_pulled = 1, pulled_at = ?,
                            pulled_by = ?, pulled_batch_id = ?, updated_at = ?,
                            version = version + 1
                        WHERE po_number IN ({placeholders})""",
                    [stamp_now, operator, batch_id, stamp_now] + pos,
                )
        conn.commit()
    finally:
        conn.close()

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream, as_attachment=True, download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/export/batches")
def api_export_batches():
    conn = get_conn()
    try:
        rows = conn.execute(
            "SELECT * FROM export_batches ORDER BY id DESC LIMIT 50"
        ).fetchall()
        return jsonify([dict(r) for r in rows])
    finally:
        conn.close()


def _same(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        # 這裡不能寫成 `str(a or "") == str(b or "")`——0 在 Python 是
        # falsy，`0 or ""` 會變成 ""，導致 None 和 0 被誤判成「一樣」。
        # 對「還沒同步過」（None）跟「同步回來是 0」（真的算出來是 0）
        # 這種一定要分得開的欄位（實際驗入數量、出貨數量…）來說，這個
        # bug 會讓資料庫的值卡死在 None，UPDATE 語句永遠不會被執行到。
        return False
    return str(a).strip() == str(b).strip()


def lan_ip():
    """抓這台電腦在辦公室網路上的位址，好告訴同事要連哪裡。"""
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # 不會真的送出封包，只是讓作業系統挑出對外那張網卡
        sock.connect(("8.8.8.8", 80))
        return sock.getsockname()[0]
    except OSError:
        try:
            return socket.gethostbyname(socket.gethostname())
        except OSError:
            return ""
    finally:
        sock.close()


if __name__ == "__main__":
    import socket as _socket

    # 5000 埠被占用最常見的原因：舊視窗沒關乾淨、背景還留著一個沒更新到
    # 的伺服器在跑。這種狀況畫面通常還打得開（連到舊的那個），怎麼換
    # 檔案都不會生效，卻完全看不出原因——所以啟動前先檢查一次，寧可
    # 明確擋下來，也不要讓兩個版本同時活著。
    probe = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
    already_running = probe.connect_ex(("127.0.0.1", 5000)) == 0
    probe.close()
    if already_running:
        print("=" * 62)
        print(" ⚠ 無法啟動：5000 這個埠已經有別的程式在用了")
        print()
        print(" 最常見的原因：還有一個舊的酷澎訂單系統視窗沒關乾淨，")
        print(" 背景還留著一份沒更新到最新版的伺服器在跑。")
        print()
        print(" 解法：")
        print(" 1. 找找看工作列或工作管理員裡，是不是還有另一個黑視窗")
        print("    （或另一個 python.exe），把它整個關掉")
        print(" 2. 打開工作管理員（Ctrl+Shift+Esc），結束所有 python.exe")
        print(" 3. 再重新雙擊 START 一次")
        print("=" * 62)
        input(" 按 Enter 關閉這個視窗...")
        raise SystemExit(1)

    moved = moved_on_start
    ip = lan_ip()

    print("=" * 62)
    print(" 酷澎訂單管理系統 已啟動")
    print(f" 版本號： {BUILD_VERSION}　（畫面右下角也會顯示這個號碼，")
    print("           兩邊對得起來，才代表你看到的是最新版）")
    print()
    print("  你自己用：      http://127.0.0.1:5000")
    if ip and ip != "127.0.0.1":
        print(f"  給同事的網址：  http://{ip}:5000")
        print("                 （同一個辦公室網路才連得到，")
        print("                   而且這台電腦要開著、程式要跑著）")
    else:
        print("  ※ 抓不到辦公室網路位址，同事可能連不進來。")
    print()
    print(f" 資料與設定放在： {db.DATA_DIR}")
    print(" （更新程式時，這個資料夾不要動，裡面是你的訂單和設定）")
    if moved:
        print()
        print(f" ※ 已把 {'、'.join(moved)} 從舊位置搬進資料資料夾")
    print("=" * 62)

    # 綁 0.0.0.0 才收得到區網連線；threaded 讓多人同時操作不用排隊
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
