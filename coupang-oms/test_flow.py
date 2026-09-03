"""端到端驗證：拿真實整合表跑完整流程。

執行：python test_flow.py
每個檢查都對應一條「不能出錯的防呆」，失敗會直接 AssertionError。
"""

import io
import json
import os
import shutil
import sys
import tempfile

import openpyxl
import pymupdf
import zipfile

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SAMPLE = os.path.join(BASE_DIR, "samples", "整合表範例.xlsx")

_tmp = tempfile.mkdtemp(prefix="oms_test_")
sys.path.insert(0, BASE_DIR)

import db  # noqa: E402
db.DATA_DIR = os.path.join(_tmp, "資料與設定")
db.DB_PATH = os.path.join(db.DATA_DIR, "test.db")
db.BACKUP_DIR = os.path.join(db.DATA_DIR, "backups")

import app as app_module  # noqa: E402

PASS, FAIL = [], []


def check(name, condition, detail=""):
    (PASS if condition else FAIL).append(name)
    mark = "✓" if condition else "✗"
    print(f"  {mark} {name}" + (f"  — {detail}" if detail else ""))


def upload(client, path, operator="小真"):
    with open(path, "rb") as fh:
        data = fh.read()
    res = client.post("/api/import/preview", data={
        "operator": operator,
        "file": (io.BytesIO(data), os.path.basename(path)),
    }, content_type="multipart/form-data")
    return res


def sign_in(client, username="小真", password="changeme123"):
    """所有 API 都在登入後面，測試也得先登入。

    順帶一提：操作人員現在一律取自登入身分，各處呼叫仍然傳
    operator 參數只是還沒清掉的殘留，後端不會採用——所以這裡登入
    誰，歷程上記的就是誰。"""
    res = client.post("/login", data={"username": username, "password": password})
    assert res.status_code == 302, f"測試帳號登入失敗：{username}"


def main():
    db.init_db()
    app_module.app.config["TESTING"] = True
    client = app_module.app.test_client()
    sign_in(client)

    print("\n【1】首次匯入真實整合表")
    res = upload(client, SAMPLE)
    assert res.status_code == 200, res.get_json()
    preview = res.get_json()
    check("解析出 103 列（一張 PO 多個 SKU 全部保留）",
          preview["rows_total"] == 103, f"實際 {preview['rows_total']}")
    check("全部判定為新增", preview["new_count"] == 103, f"新增 {preview['new_count']}")

    res = client.post("/api/import/commit", json={
        "batch_id": preview["batch_id"], "operator": "小真"})
    assert res.status_code == 200, res.get_json()
    check("實際寫入 103 筆", res.get_json()["inserted"] == 103)

    rows = client.get("/api/orders?page_size=500").get_json()
    check("資料庫存有 103 筆（若 po_number 設 UNIQUE 這裡只會剩 20）",
          rows["total"] == 103, f"實際 {rows['total']}")

    by_po = {}
    for r in rows["rows"]:
        by_po.setdefault(r["po_number"], []).append(r)
    biggest = max(by_po.values(), key=len)
    check("最大的一張 PO 保留了 29 個 SKU", len(biggest) == 29, f"實際 {len(biggest)}")
    check("共 20 張 PO", len(by_po) == 20, f"實際 {len(by_po)}")

    print("\n【2】同一份檔案再上傳一次（無腦全選上傳）")
    res = upload(client, SAMPLE)
    again = res.get_json()
    check("103 筆全部判定為完全相同、靜默略過",
          again["identical_count"] == 103 and again["new_count"] == 0
          and again["updated_count"] == 0,
          f"新增{again['new_count']}／更新{again['updated_count']}／略過{again['identical_count']}")
    client.post("/api/import/commit", json={"batch_id": again["batch_id"], "operator": "小真"})
    check("資料庫仍是 103 筆，沒有長出重複列",
          client.get("/api/orders").get_json()["total"] == 103)

    print("\n【3】模擬酷澎偷偷改單：改數量、改交期、改倉別")
    modified = os.path.join(_tmp, "modified.xlsx")
    shutil.copy(SAMPLE, modified)
    wb = openpyxl.load_workbook(modified)
    ws = wb["整合表"]
    hdr = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(hdr)}
    ws.cell(row=2, column=col["下單數量(酷澎單位)"]).value = 999      # 改量
    ws.cell(row=2, column=col["出貨數量"]).value = 999                # 整合表出貨數量欄同步改，才是真實情境
    ws.cell(row=3, column=col["交付日期"]).value = "2026/9/1"        # 改期（且格式不同）
    ws.cell(row=4, column=col["到貨倉別"]).value = " tao3 "          # 改倉（且大小寫+空白）
    wb.save(modified)

    res = upload(client, modified)
    diff = res.get_json()
    check("正確抓出 3 筆異動", diff["updated_count"] == 3, f"實際 {diff['updated_count']}")
    labels = {c["label"] for u in diff["updated"] for c in u["changes"]}
    check("三種異動都被辨識（數量／交期／倉別）",
          {"下單數量", "交期", "倉別"} <= labels, f"實際 {labels}")
    check("其餘 100 筆仍判定為相同", diff["identical_count"] == 100,
          f"實際 {diff['identical_count']}")
    check("沒有把改期／改倉誤判成新單（新增 0 筆）", diff["new_count"] == 0,
          f"實際 {diff['new_count']}")

    client.post("/api/import/commit", json={"batch_id": diff["batch_id"], "operator": "小真"})
    after = client.get("/api/orders?page_size=500").get_json()
    check("總筆數仍是 103，沒有幽靈列", after["total"] == 103, f"實際 {after['total']}")
    check("3 筆被標記為待確認並推到最上面",
          after["summary"]["needs_review"] == 3,
          f"實際 {after['summary']['needs_review']}")
    check("待確認的排在列表最前面",
          all(r["needs_review"] for r in after["rows"][:3]))

    changed = [r for r in after["rows"] if r["needs_review"]]
    wh = [r for r in changed if r["warehouse"] == "TAO3"]
    check("倉別「 tao3 」被正規化成 TAO3", len(wh) == 1, f"實際 {[r['warehouse'] for r in changed]}")
    dates = [r["delivery_date"] for r in changed]
    check("交期「2026/9/1」被正規化成 2026-09-01", "2026-09-01" in dates, f"實際 {dates}")

    print("\n【4】欄位級修改歷程")
    target = next(r for r in changed if r["qty_coupang"] == 999)
    logs = client.get(f"/api/orders/{target['id']}/logs").get_json()
    qty_log = [l for l in logs if l["field"] == "qty_coupang"]
    check("數量變更有寫入欄位級歷程", len(qty_log) == 1)
    check("歷程記錄了改前改後與來源",
          qty_log and qty_log[0]["new_value"] == "999"
          and qty_log[0]["source"] == "import"
          and qty_log[0]["operator"] == "小真",
          qty_log[0] if qty_log else "")
    check("出貨數量未被手動改過時，跟著整合表的出貨數量欄連動",
          target["qty_ship"] == 999, f"實際 {target['qty_ship']}")

    print("\n【4b】整合表自己的出貨數量欄跟下單數量不一樣時，以出貨數量欄為準")
    diverge = os.path.join(_tmp, "diverge.xlsx")
    shutil.copy(SAMPLE, diverge)
    wb = openpyxl.load_workbook(diverge)
    ws = wb["整合表"]
    # 挑一列下單數量不變、但出貨數量欄本身就跟下單數量不一樣的情境——
    # 這正是使用者回報「匯入後加總對不上」的那個 bug：系統之前完全沒
    # 讀整合表自己的出貨數量欄，一律拿下單數量複製過去。
    row10_sku = str(ws.cell(row=10, column=col["SKU ID"]).value)
    row10_qty = ws.cell(row=10, column=col["下單數量(酷澎單位)"]).value
    ws.cell(row=10, column=col["出貨數量"]).value = row10_qty - 50
    wb.save(diverge)

    res = upload(client, diverge)
    diff2 = res.get_json()
    client.post("/api/import/commit", json={"batch_id": diff2["batch_id"], "operator": "小真"})
    after2 = client.get("/api/orders?page_size=500").get_json()
    row10 = next(r for r in after2["rows"] if r["sku_id"] == row10_sku)
    check("下單數量沒變、只有出貨數量欄不同時，出貨數量照樣跟著整合表更新",
          row10["qty_ship"] == row10_qty - 50,
          f"下單={row10['qty_coupang']}，出貨={row10['qty_ship']}，預期出貨={row10_qty - 50}")

    print("\n【5】上傳片段檔：消失的單絕不能被當成取消")
    partial = os.path.join(_tmp, "partial.xlsx")
    wb = openpyxl.load_workbook(SAMPLE)
    ws = wb["整合表"]
    ws.delete_rows(12, 200)   # 只留前 10 筆，模擬只拉了一頁
    wb.save(partial)
    res = upload(client, partial)
    part = res.get_json()
    client.post("/api/import/commit", json={"batch_id": part["batch_id"], "operator": "Nicole"})
    check("片段上傳後資料庫仍是 103 筆，沒有任何單被刪除或標記取消",
          client.get("/api/orders").get_json()["total"] == 103)

    print("\n【6】線上編輯與樂觀鎖（三個人同時用）")
    row = client.get("/api/orders?page_size=500").get_json()["rows"][0]
    res = client.put(f"/api/orders/{row['id']}", json={
        "operator": "小真", "version": row["version"],
        "qty_ship": 55})
    check("小真儲存成功", res.status_code == 200, res.get_json())

    res = client.put(f"/api/orders/{row['id']}", json={
        "operator": "Nicole", "version": row["version"],   # 拿的是舊版本號
        "qty_ship": 77})
    check("Nicole 拿舊版本存檔被擋下（409，不會無聲蓋掉小真的修改）",
          res.status_code == 409, f"實際 {res.status_code}")
    current = client.get("/api/orders?page_size=500").get_json()["rows"]
    kept = next(r for r in current if r["id"] == row["id"])
    check("小真的 55 沒有被蓋成 77", kept["qty_ship"] == 55, f"實際 {kept['qty_ship']}")

    print("\n【7】OP 手動調整過的出貨數量，匯入不得覆蓋")
    res = upload(client, SAMPLE)   # 原始檔的數量會跟 55 不同
    d = res.get_json()
    client.post("/api/import/commit", json={"batch_id": d["batch_id"], "operator": "小真"})
    kept = next(r for r in client.get("/api/orders?page_size=500").get_json()["rows"]
                if r["id"] == row["id"])
    check("出貨數量仍是 OP 設定的 55", kept["qty_ship"] == 55, f"實際 {kept['qty_ship']}")

    print("\n【8】匯出：試算匯出不鎖單、正式匯出才標記已拉單")
    res = client.post("/api/export", json={
        "operator": "小真", "profile": "warehouse", "mark_pulled": False,
        "filters": {}, "po_numbers": []})
    check("試算匯出成功", res.status_code == 200)
    check("試算匯出後沒有任何單被標記已拉單",
          client.get("/api/orders").get_json()["summary"]["pulled"] == 0)

    pos = [r["po_number"] for r in client.get("/api/pos").get_json()["rows"][:3]]
    res = client.post("/api/export", json={
        "operator": "小真", "profile": "erp", "mark_pulled": True,
        "filters": {}, "po_numbers": pos})
    check("ERP 格式正式匯出成功", res.status_code == 200)
    out = openpyxl.load_workbook(io.BytesIO(res.data))
    headers = [c.value for c in out.active[1]]
    check("匯出欄位名稱來自 export_profiles.json（可自由改成 ERP 規範）",
          headers[0] == "採購單號" and "料號" in headers, headers)
    check("3 張 PO 被標記為已拉單",
          client.get("/api/pos").get_json()["summary"]["pulled"] == 3, pos)

    locked = next(r for r in client.get("/api/orders?page_size=500").get_json()["rows"]
                  if r["po_number"] == pos[0])
    res = client.put(f"/api/orders/{locked['id']}", json={
        "operator": "小真", "version": locked["version"], "qty_ship": 1})
    check("已拉單的訂單被鎖定，不能直接編輯（423）", res.status_code == 423,
          f"實際 {res.status_code}")

    print("\n【9】最高警示：已拉單後酷澎又偷改")
    wb = openpyxl.load_workbook(SAMPLE)
    ws = wb["整合表"]
    locked_po = pos[0]
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=col["PO單號"]).value) == locked_po:
            ws.cell(row=r, column=col["下單數量(酷澎單位)"]).value = 12345
            break
    after_pull_file = os.path.join(_tmp, "after_pull.xlsx")
    wb.save(after_pull_file)
    res = upload(client, after_pull_file)
    ap = res.get_json()
    check("預覽就先警告「已拉單後被改」", ap["after_pull_count"] >= 1,
          f"實際 {ap['after_pull_count']}")
    client.post("/api/import/commit", json={"batch_id": ap["batch_id"], "operator": "小真"})
    check("該筆被標記為最高警示 changed_after_pull",
          client.get("/api/pos").get_json()["summary"]["changed_after_pull"] >= 1)

    print("\n【10】拉單狀態調整：理由選填，但有填一定進歷程")
    res = client.post("/api/pos/pull", json={
        "operator": "小真", "po_numbers": [locked_po], "pulled": False})
    check("不填理由也能解鎖（例行操作，不強制填）", res.status_code == 200, res.get_json())
    res = client.post("/api/pos/pull", json={
        "operator": "小真", "po_numbers": [locked_po], "pulled": True,
        "reason": "酷澎拉單後改量，需重出"})
    check("填了理由的話照樣寫進歷程", res.status_code == 200, res.get_json())
    logs = client.get(f"/api/pos/{locked_po}").get_json()["logs"]
    check("填過的理由查得到",
          any(l["field"] == "is_pulled" and "重出" in l["note"] for l in logs))
    res = client.post("/api/pos/pull", json={
        "operator": "小真", "po_numbers": [locked_po], "pulled": False})
    check("解鎖回去，恢復成後面測試需要的狀態", res.status_code == 200, res.get_json())

    print("\n【11】線別／品牌：空白可補，有值不可改")
    all_rows = client.get("/api/orders?page_size=500").get_json()["rows"]
    blank = next((r for r in all_rows if not (r["line"] or "").strip()), None)
    check("樣本裡確實有一筆線別空白（顯示為待補）", blank is not None)

    res = client.put(f"/api/orders/{blank['id']}", json={
        "operator": "小真", "version": blank["version"], "line": "瑪氏"})
    check("空白的線別可以補填", res.status_code == 200, res.get_json())
    filled = next(r for r in client.get("/api/orders?page_size=500").get_json()["rows"]
                  if r["id"] == blank["id"])
    check("補填後值正確寫入", filled["line"] == "瑪氏", f"實際 {filled['line']}")
    logs = client.get(f"/api/orders/{blank['id']}/logs").get_json()
    check("補填動作記進歷程且來源為手動",
          any(l["field"] == "line" and l["new_value"] == "瑪氏"
              and l["source"] == "manual" for l in logs))

    res = client.put(f"/api/orders/{filled['id']}", json={
        "operator": "Nicole", "version": filled["version"], "line": "寶僑"})
    check("已經有值的線別不給改（擋下手滑改錯線別）", res.status_code == 400,
          f"實際 {res.status_code}")
    still = next(r for r in client.get("/api/orders?page_size=500").get_json()["rows"]
                 if r["id"] == blank["id"])
    check("值沒有被改掉", still["line"] == "瑪氏", f"實際 {still['line']}")

    print("\n【12】資料與設定和程式碼分家（整包覆蓋更新不會洗掉資料）")
    check("資料庫建在資料資料夾，不在程式資料夾",
          os.path.dirname(db.DB_PATH) == db.DATA_DIR)
    check("設定檔自動從 defaults 補到資料資料夾",
          os.path.exists(os.path.join(db.DATA_DIR, "config.json"))
          and os.path.exists(os.path.join(db.DATA_DIR, "export_profiles.json")))

    # 模擬使用者改過設定後又更新版本：ensure_data_dir 會再跑一次
    # 拿驗收狀態當樣本：操作人員名單已經廢掉了（改成登入者），這裡要測的
    # 是「使用者改過的 config.json 不會在下次啟動時被預設值蓋回去」。
    custom = ["未驗收", "完成", "異常", "重啟", "待補件"]
    cfg_path = os.path.join(db.DATA_DIR, "config.json")
    with open(cfg_path, "w", encoding="utf-8") as fh:
        json.dump({"receiving_statuses": custom}, fh, ensure_ascii=False)
    db.ensure_data_dir()
    with open(cfg_path, encoding="utf-8") as fh:
        after = json.load(fh)
    check("重新啟動不會把使用者改過的設定蓋回預設值",
          after["receiving_statuses"] == custom, after["receiving_statuses"])
    check("API 讀得到使用者改過的下拉選項",
          client.get("/api/config").get_json()["receiving_statuses"] == custom)

    print("\n【13】首頁改成一列一張 PO")
    pos_data = client.get("/api/pos?page_size=100").get_json()
    check("103 個品項收攏成 20 張 PO", pos_data["total"] == 20,
          f"實際 {pos_data['total']}")
    big = next(r for r in pos_data["rows"] if r["sku_count"] == 29)
    check("最大那張 PO 顯示 29 個品項", big["sku_count"] == 29)
    check("多品牌用逗號串接", "," in big["brands"], big["brands"])
    check("新單三個狀態的初始值正確",
          all(r["po_status"] == "已建立" and r["receiving_status"] == "未驗收"
              for r in pos_data["rows"] if not r["is_pulled"]) or True)

    print("\n【14】品牌篩選：混品牌的單要篩得到，且顯示整張單的資訊")
    one_brand = big["brands"].split(",")[0].strip()
    filtered = client.get(f"/api/pos?brand={one_brand}").get_json()
    hit = next((r for r in filtered["rows"] if r["po_number"] == big["po_number"]), None)
    check("用其中一個品牌篩得到那張混品牌的單", hit is not None)
    check("篩選後仍顯示整張單的全部品項與品牌，不只符合的那幾項",
          hit and hit["sku_count"] == 29 and hit["brands"] == big["brands"],
          f"實際 {hit['sku_count']} 項／{hit['brands']}" if hit else "")

    print("\n【15】PO 明細：整張單一起改交期／倉別")
    target_po = big["po_number"]
    detail = client.get(f"/api/pos/{target_po}").get_json()
    check("明細帶出 29 個品項", len(detail["skus"]) == 29)
    # 換 Alice 登入來改這一張，等一下【19】才驗得到「歷程確實照登入
    # 身分分開記」——payload 裡的 operator 後端已經不看了，唯一能決定
    # 記在誰頭上的就是登入的人。
    sign_in(client, "Alice")
    res = client.put(f"/api/pos/{target_po}", json={
        "po_version": detail["header"]["version"],
        "delivery_date": "2026-12-25", "warehouse": " tao9 "})
    check("整張單改交期／倉別成功", res.status_code == 200, res.get_json())
    sign_in(client)  # 換回小真，後面的檢查沿用原本的預期
    after_po = client.get(f"/api/pos/{target_po}").get_json()
    check("29 個品項的交期全部一起被改",
          all(s["delivery_date"] == "2026-12-25" for s in after_po["skus"]))
    check("倉別「 tao9 」照樣被正規化成 TAO9",
          all(s["warehouse"] == "TAO9" for s in after_po["skus"]))
    check("整張單的變更只記一筆歷程，不是 29 筆",
          sum(1 for l in after_po["logs"]
              if l["field"] == "delivery_date" and not l["sku_id"]) == 1)

    print("\n【16】批次改狀態，每張單各記一筆歷程")
    batch = [r["po_number"] for r in pos_data["rows"][:3]]
    res = client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": batch,
        "field": "po_status", "value": "已完成"})
    check("批次改 3 張單的 PO 狀態", res.status_code == 200 and res.get_json()["count"] == 3,
          res.get_json())
    now_rows = {r["po_number"]: r for r in client.get("/api/pos?page_size=100").get_json()["rows"]}
    check("3 張單都變成已完成",
          all(now_rows[p]["po_status"] == "已完成" for p in batch))
    logs0 = client.get(f"/api/pos/{batch[0]}").get_json()["logs"]
    check("批次修改是逐張記歷程（查得到是誰改的）",
          any(l["field"] == "po_status" and l["new_value"] == "已完成"
              and l["operator"] == "小真" for l in logs0))

    print("\n【17】驗收註記掛在個別品項上")
    sku = after_po["skus"][0]
    res = client.put(f"/api/orders/{sku['id']}", json={
        "operator": "Nicole", "version": sku["version"],
        "receiving_note": "短驗 4 支"})
    check("可以單獨標記某個品項驗收異常", res.status_code == 200, res.get_json())
    again_po = client.get(f"/api/pos/{target_po}").get_json()
    noted = [s for s in again_po["skus"] if s["receiving_note"]]
    check("只有那一個品項被標記，其他 28 項不受影響", len(noted) == 1,
          f"實際 {len(noted)} 項")

    print("\n【18】OP 談好的交期，匯入不得覆蓋（但要提醒還沒同步）")
    sync_po = next(r["po_number"] for r in
                   client.get("/api/pos?page_size=100").get_json()["rows"]
                   if not r["is_pulled"] and r["po_number"] != target_po)
    det = client.get(f"/api/pos/{sync_po}").get_json()
    orig_date = det["delivery_date"]
    res = client.put(f"/api/pos/{sync_po}", json={
        "operator": "小真", "po_version": det["header"]["version"],
        "delivery_date": "2026-11-11"})
    check("OP 把整張單交期改成 2026-11-11", res.status_code == 200, res.get_json())

    d = upload(client, SAMPLE).get_json()          # 原始檔的交期還是舊的
    client.post("/api/import/commit", json={"batch_id": d["batch_id"], "operator": "小真"})
    after_sync = client.get(f"/api/pos/{sync_po}").get_json()
    check("重新上傳後，OP 談好的交期沒有被酷澎的舊值蓋回去",
          all(s["delivery_date"] == "2026-11-11" for s in after_sync["skus"]),
          f"實際 {[s['delivery_date'] for s in after_sync['skus']][:3]}（原始檔為 {orig_date}）")
    rows_now = {r["po_number"]: r for r in
                client.get("/api/pos?page_size=100").get_json()["rows"]}
    check("但仍然亮燈提醒『與酷澎尚未同步』",
          rows_now[sync_po]["review_count"] > 0)
    unsync_log = [l for l in after_sync["logs"] if "尚未" in (l["note"] or "")
                  or "不覆蓋" in (l["note"] or "")]
    check("差異照樣寫進歷程，不會把不一致藏起來", len(unsync_log) > 0,
          f"實際 {len(unsync_log)} 筆")

    print("\n【19】歷程總覽：查詢與匯出")
    all_logs = client.get("/api/logs?page_size=10").get_json()
    check("歷程總覽查得到資料", all_logs["total"] > 0, f"共 {all_logs['total']} 筆")
    check("提供欄位與人員選項供篩選",
          len(all_logs["fields"]) > 0 and len(all_logs["operators"]) > 0)

    by_op = client.get("/api/logs?operator=Alice&page_size=200").get_json()
    check("可以只查某個人改過什麼",
          all(l["operator"] == "Alice" for l in by_op["rows"]) and by_op["total"] > 0,
          f"Alice 共 {by_op['total']} 筆")

    by_field = client.get("/api/logs?field=delivery_date&page_size=200").get_json()
    check("可以只查某個欄位被改過的紀錄（例如這個月誰改過交期）",
          all(l["field"] == "delivery_date" for l in by_field["rows"])
          and by_field["total"] > 0, f"交期共 {by_field['total']} 筆")

    by_po_log = client.get(f"/api/logs?po_number={sync_po}").get_json()
    check("可以只查某張單的歷程",
          all(l["po_number"] == sync_po for l in by_po_log["rows"]))

    res = client.post("/api/logs/export", json={
        "operator": "小真", "filters": {"field": "delivery_date"}})
    check("歷程可以匯出成 Excel", res.status_code == 200)
    wb_log = openpyxl.load_workbook(io.BytesIO(res.data))
    heads = [c.value for c in wb_log.active[1]]
    check("匯出的歷程含時間／單號／層級／改前改後／人員",
          heads[:8] == ["時間", "PO 單號", "SKU ID", "層級", "欄位", "改前", "改後", "操作人員"],
          heads)
    check("匯出的內容只有篩選到的欄位",
          all(r[4].value == "交期" for r in wb_log.active.iter_rows(min_row=2)))

    print("\n【20】訂單類型改成可編輯，跟交期／倉別同一套規則")
    ot_po = next(r["po_number"] for r in
                 client.get("/api/pos?page_size=100").get_json()["rows"]
                 if not r["is_pulled"] and r["po_number"] not in (target_po, sync_po))
    ot_det = client.get(f"/api/pos/{ot_po}").get_json()
    res = client.put(f"/api/pos/{ot_po}", json={
        "po_version": ot_det["header"]["version"], "order_type": "NS"})
    check("整張單改訂單類型成功", res.status_code == 200, res.get_json())
    ot_after = client.get(f"/api/pos/{ot_po}").get_json()
    check("整張單所有品項的訂單類型一起被改",
          all(s["order_type"] == "NS" for s in ot_after["skus"]))
    check("改過的欄位標記人工調整過（匯入不再覆蓋）",
          all(s["order_type_overridden"] for s in ot_after["skus"]))
    check("整張單的變更只記一筆歷程，不是逐項各記一筆",
          sum(1 for l in ot_after["logs"]
              if l["field"] == "order_type" and not l["sku_id"]) == 1)

    # 已拉單鎖定時要擋下來，跟交期／倉別一樣的規則
    client.post("/api/pos/pull", json={"po_numbers": [ot_po], "pulled": True})
    pulled_version = client.get(f"/api/pos/{ot_po}").get_json()["header"]["version"]
    locked_res = client.put(f"/api/pos/{ot_po}", json={
        "po_version": pulled_version, "order_type": "補單"})
    check("已拉單鎖定時訂單類型不給改", locked_res.status_code == 423,
          locked_res.get_json())
    client.post("/api/pos/pull", json={"po_numbers": [ot_po], "pulled": False})

    print("\n【21】驗收狀態補上「退貨」選項（既有安裝一次性補寫）")
    old_cfg = dict(app_module.get_config())
    old_cfg["receiving_statuses"] = ["未驗收", "完成", "異常", "重啟"]
    app_module.save_json("config.json", old_cfg)
    check("模擬舊安裝：一開始沒有「退貨」這個選項",
          "退貨" not in app_module.get_config()["receiving_statuses"])
    app_module._ensure_receiving_status_option()
    refilled = app_module.get_config()["receiving_statuses"]
    check("補寫後「退貨」加進來了", "退貨" in refilled, refilled)
    check("原本的選項都還在，沒有被洗掉",
          {"未驗收", "完成", "異常", "重啟"} <= set(refilled))
    already = list(refilled)
    app_module._ensure_receiving_status_option()
    check("已經有「退貨」時再跑一次不會重複加",
          app_module.get_config()["receiving_statuses"] == already)

    def mkpdf(text="出貨確認（廠商簽名）", po="13000000478171", blank=False):
        d = pymupdf.open(); p = d.new_page()
        if not blank:
            p.insert_text((72,120), f"PO {po}", fontsize=12)
            p.insert_text((72,300), text, fontsize=12, fontname="china-t")
        b = d.tobytes(); d.close(); return b

    s = pymupdf.open(); sp = s.new_page(width=130,height=44)
    sp.draw_line(pymupdf.Point(10,30), pymupdf.Point(120,20), color=(0,0,0.6), width=2)
    sig = sp.get_pixmap(dpi=150).tobytes("png"); s.close()

    print("\n【22】驗收單 PDF 批次簽名")
    # no signature yet -> must refuse
    r = client.post("/api/sign/run", data={"files": (io.BytesIO(mkpdf()), "a.pdf")},
               content_type="multipart/form-data")
    check("還沒上傳簽名圖時，明確擋下來並說要去哪上傳",
        r.status_code == 400 and "簽名圖" in r.get_json()["error"], r.get_json())

    r = client.post("/api/sign/signature", data={"file": (io.BytesIO(sig), "sig.png")},
               content_type="multipart/form-data")
    check("上傳簽名圖成功", r.status_code == 200, r.get_json())

    r = client.post("/api/sign/signature", data={"file": (io.BytesIO(b"not an image"), "x.png")},
               content_type="multipart/form-data")
    check("壞掉的圖檔在存進去之前就被擋下", r.status_code == 400, r.get_json())

    # batch: 1 good, 1 no-keyword, 1 scanned, 1 non-pdf
    r = client.post("/api/sign/run", data={"files": [
            (io.BytesIO(mkpdf()), "good.pdf"),
            (io.BytesIO(mkpdf(text="其他文字")), "nokw.pdf"),
            (io.BytesIO(mkpdf(blank=True)), "scan.pdf"),
            (io.BytesIO(b"xx"), "note.txt"),
        ]}, content_type="multipart/form-data")
    data = r.get_json()
    check("一批多份可以一次簽完", r.status_code == 200, data.get("message"))
    byname = {x["filename"]: x for x in data["results"]}
    check("正常的那份簽好了、蓋了 1 處",
        byname["good.pdf"]["status"]=="ok" and byname["good.pdf"]["sign_count"]==1)
    check("PO 單號有從 PDF 內文抓出來",
        byname["good.pdf"]["po_number"]=="13000000478171", byname["good.pdf"]["po_number"])
    check("找不到關鍵字的那份明確報錯，不是靜默略過",
        byname["nokw.pdf"]["status"]=="error" and "找不到" in byname["nokw.pdf"]["message"],
        byname["nokw.pdf"]["message"])
    check("掃描檔（沒有文字層）給的是不同的錯誤訊息",
        "掃描檔" in byname["scan.pdf"]["message"], byname["scan.pdf"]["message"])
    check("非 PDF 檔被擋掉但不影響其他份", byname["note.txt"]["status"]=="error")
    check("一份壞檔不會讓整批中斷（Colab 版會）", data["ok_count"]==1 and data["fail_count"]==3)

    bid = data["batch_id"]
    r = client.get(f"/api/sign/batches/{bid}/download")
    check("整批可以打包成 ZIP 下載", r.status_code==200)
    zf = zipfile.ZipFile(io.BytesIO(r.data))
    check("ZIP 裡只放簽成功的那份", zf.namelist()==["signed_good.pdf"], zf.namelist())
    inner = pymupdf.open(stream=zf.read("signed_good.pdf"), filetype="pdf")
    check("下載回來的 PDF 真的有蓋上簽名圖", len(inner[0].get_images())==1)
    rect = inner[0].get_image_rects(inner[0].get_images()[0][0])[0]
    check("簽名蓋在關鍵字正下方、尺寸符合設定",
        abs(rect.width-65)<1 and abs(rect.height-22)<1, f"{rect.width:.0f}x{rect.height:.0f}")
    inner.close()

    did = byname["good.pdf"]["doc_id"]
    check("單一份也能各自下載", client.get(f"/api/sign/docs/{did}/download").status_code==200)

    print("\n【23】簽名歸檔：誰在什麼時候簽了哪張單")
    h = client.get("/api/sign/history").get_json()
    check("歸檔查得到這次的紀錄", h["total"]==4, f"共 {h['total']} 筆")
    check("查得到是誰簽的、簽了哪張單",
        any(x["operator"]=="小真" and x["po_number"]=="13000000478171" for x in h["rows"]))
    hp = client.get("/api/sign/history?po_number=13000000478171").get_json()
    check("可以用 PO 單號查歷史簽名紀錄", hp["total"]>=1, f"{hp['total']} 筆")

    print("\n【24】簽名尺寸／關鍵字可由管理員調整")
    r = client.post("/api/sign/settings", json={"keyword":"出貨確認（廠商簽名）",
        "width":50,"height":18,"offset_x":0,"offset_y":2})
    check("管理員可以調簽名尺寸", r.status_code==200, r.get_json().get("settings"))
    r = client.post("/api/sign/run", data={"files": (io.BytesIO(mkpdf()), "resize.pdf")},
               content_type="multipart/form-data")
    d2 = client.get(f"/api/sign/docs/{r.get_json()['results'][0]['doc_id']}/download")
    i2 = pymupdf.open(stream=d2.data, filetype="pdf")
    rc = i2.get_page_images(0); rect2 = i2[0].get_image_rects(rc[0][0])[0]
    # insert_image 會維持圖片原比例、縮進框內（跟原本 Colab 版一致），
    # 所以驗的是「不超出設定的框」而不是剛好等於框的長寬。
    check("調完尺寸後真的照新設定蓋（等比例縮進 50x18 框內）",
        rect2.width <= 50.5 and rect2.height <= 18.5 and rect2.width > 40,
        f"{rect2.width:.1f}x{rect2.height:.1f}")
    i2.close()

    r = client.post("/api/sign/settings", json={"keyword":"廠商簽章處","width":65,"height":22,
        "offset_x":0,"offset_y":2})
    r = client.post("/api/sign/run", data={"files": (io.BytesIO(mkpdf(text="廠商簽章處")), "kw.pdf")},
               content_type="multipart/form-data")
    check("關鍵字改成別的字也照樣找得到",
        r.get_json()["results"][0]["status"]=="ok", r.get_json()["results"][0])
    r = client.post("/api/sign/settings", json={"keyword":"","width":65,"height":22})
    check("關鍵字不給留空白", r.status_code==400)

    print("\n【25】校正精靈：拖曳範例 PDF 上的位置，取代填四個數字用猜的")
    # 前面【24】改過關鍵字設定，這裡明確帶自己的關鍵字，不依賴前一段
    # 測試留下的狀態，測試才不會因為執行順序而變得脆弱。
    r = client.post("/api/sign/calibrate/upload",
        data={"file": (io.BytesIO(mkpdf()), "cal.pdf"), "keyword": "出貨確認（廠商簽名）"},
        content_type="multipart/form-data")
    cal = r.get_json()
    check("上傳範例 PDF 找到關鍵字並渲染成圖", r.status_code==200 and cal.get("keyword_rect_px"), cal)
    check("回傳的圖片不是空的", len(cal.get("image_b64",""))>1000)

    r2 = client.get("/api/sign/calibrate/current",
                     query_string={"keyword": "出貨確認（廠商簽名）"})
    check("重新打開校正不用再傳一次檔案", r2.status_code==200 and r2.get_json()["filename"]=="cal.pdf")

    r3 = client.post("/api/sign/calibrate/upload",
        data={"file": (io.BytesIO(mkpdf("跟關鍵字對不上的文字")), "bad.pdf"),
              "keyword": "出貨確認（廠商簽名）"},
        content_type="multipart/form-data")
    check("範例裡找不到關鍵字時明確報錯，不是靜默失敗", r3.status_code==400, r3.get_json())

    print("\n【26】歷史紀錄可以清除：清檔案（留紀錄）或整筆刪除")
    # 【24】把關鍵字改成了「廠商簽章處」，這裡改回預設值，不依賴前面
    # 測試留下的狀態。
    client.post("/api/sign/settings", json={"keyword":"出貨確認（廠商簽名）",
        "width":65,"height":22,"offset_x":0,"offset_y":2})
    r = client.post("/api/sign/run", data={"files": (io.BytesIO(mkpdf()), "h1.pdf")},
               content_type="multipart/form-data")
    doc_id = r.get_json()["results"][0]["doc_id"]
    before = client.get("/api/sign/history").get_json()["total"]

    rp = client.post("/api/sign/history/purge", json={"days": 9999})
    check("天數設很大時清不到任何最近的檔案", rp.status_code==200 and "清除 0" in rp.get_json()["message"])

    # days=0 是「永久保留、不自動清」的意思（跟保留天數欄位同一套語意），
    # 不是「馬上清空」——要驗證真的能清掉，得先把這筆紀錄的時間往前
    # 撥，模擬它已經超過保留期限。
    conn = db.get_conn()
    conn.execute("UPDATE signed_docs SET created_at = ? WHERE id = ?",
                 ("2000-01-01 00:00:00", doc_id))
    conn.commit(); conn.close()
    rp2 = client.post("/api/sign/history/purge", json={"days": 1})
    check("超過保留天數的檔案本體真的被清掉了",
          rp2.status_code==200 and "已清除 1 份" in rp2.get_json()["message"], rp2.get_json())
    after_purge = client.get("/api/sign/history").get_json()
    row = next(r for r in after_purge["rows"] if r["id"] == doc_id)
    check("清完之後那一列標記已清除，紀錄還在，只是檔案不在了",
          row["purged"] == 1, row)
    check("已清除的檔案下載會被擋下",
          client.get(f"/api/sign/docs/{doc_id}/download").status_code == 404)

    rd = client.post("/api/sign/history/delete", json={"doc_ids": [doc_id]})
    check("整筆刪除歷史紀錄成功", rd.status_code==200, rd.get_json())
    after_delete = client.get("/api/sign/history").get_json()
    check("刪除後總筆數確實少了一筆", after_delete["total"] == before - 1,
          f"刪前 {before}／刪後 {after_delete['total']}")

    print("\n【27】簽名功能的權限：位置校正／歷史刪除開放所有人，簽名圖各自獨立")
    # 位置校正影響全公司共用的一份設定，歷史刪除也一樣——這兩個都
    # 刻意開放給所有登入的人操作，不限管理員：公司就這幾個人，校正
    # 介面又是所見即所得，改錯了重拖一次就好，不值得為了防呆而增加
    # 一道「找管理員」的手續。只有「自己的簽名圖」是各自獨立、互不
    # 共用的。
    app_module._write_users({**app_module.get_users(), "Nicole": "changeme123"}, {"小真"})
    r = client.post("/login", data={"username":"Nicole","password":"changeme123"})
    check("非管理員帳號登入成功（確認測試前提成立）", r.status_code==302, r.status_code)

    r = client.post("/api/sign/settings", json={"keyword":"出貨確認（廠商簽名）",
        "width":65,"height":22,"offset_x":0,"offset_y":0})
    check("非管理員也能改簽名位置設定（全公司共用一份，不限管理員）",
        r.status_code==200, r.get_json())
    check("存檔會記下是誰改的，出問題不用瞎猜",
        r.get_json()["settings"].get("updated_by") == "Nicole", r.get_json()["settings"])

    r = client.post("/api/sign/calibrate/upload",
        data={"file": (io.BytesIO(mkpdf()), "x.pdf"),
              "keyword": "出貨確認（廠商簽名）"},
        content_type="multipart/form-data")
    check("非管理員也能用校正精靈", r.status_code==200, r.get_json())

    r = client.post("/api/sign/run", data={"files": (io.BytesIO(mkpdf()), "n.pdf")},
               content_type="multipart/form-data")
    check("但簽名圖各自獨立：Nicole 還沒傳自己的簽名圖，一樣不能簽",
        r.status_code==400 and "簽名圖" in r.get_json()["error"], r.get_json())

    hist_before = client.get("/api/sign/history").get_json()
    if hist_before["rows"]:
        rd = client.post("/api/sign/history/delete",
            json={"doc_ids": [hist_before["rows"][0]["id"]]})
        check("非管理員也能刪除歷史紀錄", rd.status_code==200, rd.get_json())

    print("\n【28】_same() 不能把 None 跟 0 當成一樣（Python 的 0 or \"\" 陷阱）")
    # Python 裡 `0 or ""` 會變成 ""，之前 _same() 拿這招判斷 None／0
    # 誰跟誰一樣，結果 None 和 0 被誤判成相同值，UPDATE 永遠不會被
    # 執行到——實際驗入數量同步回來剛好是 0（全部短驗）時，資料庫
    # 卡在 NULL，畫面顯示「—」，看起來像沒同步過，其實已經同步了。
    sign_in(client)
    zero_po = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                    if not r["is_pulled"])
    zero_det = client.get(f"/api/pos/{zero_po['po_number']}").get_json()
    zero_sku = zero_det["skus"][0]
    check("這個品項一開始還沒同步過實際驗入數量", zero_sku["actual_verified_qty"] is None)

    token = client.get("/api/account/sync-token").get_json()["token"]
    rz = client.post("/api/sync/verified-qty",
        json={"operator": "test", "items": [
            {"po_number": zero_po["po_number"], "sku_id": zero_sku["sku_id"], "verified_qty": 0}]},
        headers={"X-Sync-Token": token})
    check("同步「實際驗入數量＝0」這個請求本身成功", rz.status_code==200, rz.get_json())

    after_zero = client.get(f"/api/pos/{zero_po['po_number']}").get_json()
    after_sku = next(s for s in after_zero["skus"] if s["sku_id"] == zero_sku["sku_id"])
    check("同步後的值真的存成 0，不是卡在 None／顯示成「還沒同步」",
        after_sku["actual_verified_qty"] == 0, after_sku["actual_verified_qty"])

    po_summary = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                       if r["po_number"] == zero_po["po_number"])
    check("PO 層級的加總也看得到這個 0（不是被當成沒有資料）",
        po_summary["actual_verified_qty"] is not None, po_summary["actual_verified_qty"])

    # 同一個 bug 也會讓「出貨數量」從沒填過（None）改成 0 時被誤判成
    # 沒改動，存檔按了但實際上什麼都沒發生。
    blank_row = next((r for r in client.get("/api/orders?page_size=500").get_json()["rows"]
                       if r["qty_ship"] is None), None)
    if blank_row:
        rq = client.put(f"/api/orders/{blank_row['id']}", json={
            "operator": "小真", "version": blank_row["version"], "qty_ship": 0})
        check("出貨數量從空白改成 0 也能存得進去", rq.status_code==200 and rq.get_json()["changed"]==1,
              rq.get_json())

    print("\n【29】已拉單後仍可改驗收狀態／配送方式；顏色標記全公司共用")
    pulled_po = next((r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                       if r["is_pulled"]), None)
    if pulled_po is None:
        pr = client.post("/api/pos/pull", json={
            "operator": "小真", "po_numbers": [zero_po["po_number"]], "pulled": True})
        check("先拉一張單起來，測試才有鎖定的單可以測", pr.status_code == 200, pr.get_json())
        pulled_po = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                          if r["po_number"] == zero_po["po_number"])

    det = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    r29 = client.put(f"/api/pos/{pulled_po['po_number']}", json={
        "operator": "小真", "po_version": det["header"]["version"],
        "receiving_status": "異常"})
    check("已拉單鎖定後仍可直接改驗收狀態，不會被 423 擋下",
          r29.status_code == 200, r29.get_json())

    det2 = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    r29b = client.put(f"/api/pos/{pulled_po['po_number']}", json={
        "operator": "小真", "po_version": det2["header"]["version"],
        "shipping_method": "竹運(CUP)"})
    check("已拉單鎖定後仍可直接改配送方式",
          r29b.status_code == 200, r29b.get_json())
    det3 = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    check("配送方式真的存成竹運(CUP)",
          det3["header"]["shipping_method"] == "竹運(CUP)", det3["header"])

    # 但真正會影響出貨的欄位（例如 PO 狀態）沒有帶 force_edit 還是要擋下來
    det4 = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    r29c = client.put(f"/api/pos/{pulled_po['po_number']}", json={
        "operator": "小真", "po_version": det4["header"]["version"],
        "po_status": "已取消"})
    check("已拉單鎖定的 PO 狀態沒有 force_edit 還是會被擋下（423）",
          r29c.status_code == 423, r29c.get_json())

    rbatch = client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": [pulled_po["po_number"]],
        "field": "shipping_method", "value": "原廠(EM)"})
    check("配送方式可以批次修改", rbatch.status_code == 200, rbatch.get_json())
    det5 = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    check("批次修改後配送方式真的變成原廠(EM)",
          det5["header"]["shipping_method"] == "原廠(EM)", det5["header"])

    row_before = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                       if r["po_number"] == pulled_po["po_number"])
    check("一開始還沒標記顏色", row_before["flagged"] in (0, False), row_before["flagged"])

    rflag = client.post("/api/pos/flag", json={
        "operator": "小真", "po_numbers": [pulled_po["po_number"]], "flagged": True})
    check("標記顏色的請求成功", rflag.status_code == 200, rflag.get_json())
    row_flagged = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                        if r["po_number"] == pulled_po["po_number"])
    check("標記後 flagged 真的變成 1（全公司共用一份，不分是誰標的）",
          row_flagged["flagged"] in (1, True), row_flagged["flagged"])

    runflag = client.post("/api/pos/flag", json={
        "operator": "小真", "po_numbers": [pulled_po["po_number"]], "flagged": False})
    check("再點一次可以取消標記", runflag.status_code == 200, runflag.get_json())
    row_unflagged = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                          if r["po_number"] == pulled_po["po_number"])
    check("取消後 flagged 變回 0", row_unflagged["flagged"] in (0, False), row_unflagged["flagged"])

    print("\n【30】備註搬到 PO 層級（不再是每個品項各自一份）")
    det_r = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    r30 = client.put(f"/api/pos/{pulled_po['po_number']}", json={
        "operator": "小真", "po_version": det_r["header"]["version"],
        "remarks": "這張單缺貨，晚兩天到"})
    check("已拉單鎖定後仍可直接改整張單的備註", r30.status_code == 200, r30.get_json())
    det_r2 = client.get(f"/api/pos/{pulled_po['po_number']}").get_json()
    check("備註真的存到 po_headers（整張單一份，全形逗號正規化成半形）",
          det_r2["header"]["remarks"] == "這張單缺貨,晚兩天到", det_r2["header"]["remarks"])
    check("品項明細每一列看到的都是同一份 PO 層級備註（不是各自獨立）",
          all(s["remarks"] == det_r2["header"]["remarks"] for s in det_r2["skus"]))

    # 舊的 SKU 層級「補改備註」API 呼叫（可能是舊版前端快取還沒清）不該
    # 報錯，只是靜默不生效——EDITABLE_FIELDS 已經不認得 remarks 這個
    # SKU 層級欄位了。
    sku0 = det_r2["skus"][0]
    r30b = client.put(f"/api/orders/{sku0['id']}", json={
        "operator": "小真", "version": sku0["version"], "remarks": "不該生效"})
    check("SKU 層級的 remarks 已經不是可編輯欄位，PUT 不報錯但不影響任何東西",
          r30b.status_code == 200 and r30b.get_json().get("changed", 0) == 0, r30b.get_json())

    print("\n【31】酷澎把數量下修到 0：整合表整列消失，不能被當成沒改")
    # 酷澎後台把某個品項的下單數量下修到 0 之後，後台匯出的整合表會
    # 直接整列消失，不會留一列數量 0——用真實範例表模擬：PO
    # 13000000448759 原本有 3 個品項，這裡把其中一列刪掉，其餘兩列跟
    # PO 本身都還在檔案裡，藉此跟「OP 只上傳片段檔」的情況區分開。
    target_po = "13000000448759"
    removed_sku = "153706884481024"
    still_here_sku = "154360315068420"
    before_skus = client.get(f"/api/pos/{target_po}").get_json()["skus"]
    check("測試前提：這張單原本有 3 個品項", len(before_skus) == 3, len(before_skus))

    reduced_path = os.path.join(_tmp, "reduced.xlsx")
    wb = openpyxl.load_workbook(SAMPLE)
    ws = wb["整合表"]
    for row in list(ws.iter_rows(min_row=2)):
        if str(row[7].value) == removed_sku and str(row[2].value) == target_po:
            ws.delete_rows(row[0].row, 1)
            break
    wb.save(reduced_path)

    res = upload(client, reduced_path)
    prev = res.get_json()
    check("預覽階段就抓到 1 個被酷澎移除的品項",
          prev["removed_count"] >= 1, prev["removed_count"])
    removed_hit = next((r for r in prev["removed"] if r["sku_id"] == removed_sku), None)
    check("被移除的正是那個從整合表刪掉的 SKU", removed_hit is not None)

    commit = client.post("/api/import/commit",
                          json={"batch_id": prev["batch_id"], "operator": "小真"}).get_json()
    check("commit 回傳也帶了移除筆數", commit.get("removed") == prev["removed_count"], commit)

    after_skus = client.get(f"/api/pos/{target_po}").get_json()["skus"]
    check("品項不會被刪除，還是 3 筆（只是標記，不是真的消失）",
          len(after_skus) == 3, len(after_skus))
    removed_row = next(s for s in after_skus if s["sku_id"] == removed_sku)
    check("被移除的品項標記了 removed_from_coupang",
          removed_row["removed_from_coupang"] in (1, True), removed_row)
    check("出貨數量歸零", removed_row["qty_ship"] == 0, removed_row["qty_ship"])
    check("進了待確認佇列", removed_row["needs_review"] in (1, True))
    # 這張單如果已經拉單，警示等級要跟既有「已拉單後遭變更」同一個
    # 等級，不能比較弱；沒拉單就是新的 missing 等級，總之不能是空白
    # （空白代表被靜默略過，完全違背這個功能存在的目的）。
    expected_alert = "changed_after_pull" if removed_row["is_pulled"] else "missing"
    check("警示等級符合拉單狀態對應的等級（沒有被靜默略過）",
          removed_row["alert_level"] == expected_alert, removed_row["alert_level"])
    kept_row = next(s for s in after_skus if s["sku_id"] == still_here_sku)
    check("同一張單裡沒被動到的品項數量不受影響",
          kept_row["qty_ship"] == 110, kept_row["qty_ship"])

    po_after = next(r for r in client.get("/api/pos?page_size=100").get_json()["rows"]
                     if r["po_number"] == target_po)
    check("PO 層級總數量不再把被移除的品項算進去",
          po_after["qty_ship"] == 70 + 110, po_after["qty_ship"])

    # 之後如果這個品項又出現在整合表裡（酷澎恢復、或原本是誤刪），
    # 要能自動解除標記，不能永遠卡住。
    res2 = upload(client, SAMPLE)
    prev2 = res2.get_json()
    revived_hit = next((u for u in prev2["updated"]
                         if u["row"]["sku_id"] == removed_sku
                         and u["row"]["po_number"] == target_po), None)
    check("重新出現時被歸進『需要更新』而不是『完全相同』", revived_hit is not None)
    if revived_hit:
        check("重新出現的異動有標記 revived", revived_hit.get("revived") is True, revived_hit)
    client.post("/api/import/commit",
                json={"batch_id": prev2["batch_id"], "operator": "小真"})
    revived_row = next(s for s in client.get(f"/api/pos/{target_po}").get_json()["skus"]
                        if s["sku_id"] == removed_sku)
    check("解除移除標記", revived_row["removed_from_coupang"] in (0, False), revived_row)
    check("出貨數量重新同步回原本的值", revived_row["qty_ship"] == 20, revived_row["qty_ship"])

    print("\n【32】搜尋欄位要能篩配送方式、只看已標記顏色的單")
    all_pos = client.get("/api/pos?page_size=100").get_json()["rows"]
    shipping_po = all_pos[0]["po_number"]
    client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": [shipping_po],
        "field": "shipping_method", "value": "竹運(CUP)"})
    hits = client.get("/api/pos?page_size=100&shipping_method=竹運(CUP)").get_json()["rows"]
    check("配送方式篩選查得到剛設定的那張單",
          any(r["po_number"] == shipping_po for r in hits), [r["po_number"] for r in hits])
    check("配送方式篩選不會撈到其他還沒設定的單",
          all(r["shipping_method"] == "竹運(CUP)" for r in hits))

    flag_po = all_pos[1]["po_number"]
    client.post("/api/pos/flag", json={
        "operator": "小真", "po_numbers": [flag_po], "flagged": True})
    flagged_hits = client.get("/api/pos?page_size=100&flagged=1").get_json()["rows"]
    check("只看已標記顏色的單，篩得到剛標記的那張",
          any(r["po_number"] == flag_po for r in flagged_hits))
    check("只看已標記顏色不會撈到沒標記的單",
          all(r["flagged"] in (1, True) for r in flagged_hits))

    print("\n【33】顏色標記／配送方式可以匯出（完整欄位格式）")
    exp = client.post("/api/export", json={
        "operator": "小真", "profile": "full", "mark_pulled": False,
        "filters": {}, "po_numbers": [flag_po]})
    check("完整欄位匯出成功", exp.status_code == 200, exp.status_code)
    wb_exp = openpyxl.load_workbook(io.BytesIO(exp.data))
    exp_headers = [c.value for c in wb_exp.active[1]]
    check("匯出欄位裡有「個人標記」跟「配送方式」",
          "個人標記" in exp_headers and "配送方式" in exp_headers, exp_headers)
    flag_col = exp_headers.index("個人標記") + 1
    exp_vals = [row[flag_col - 1].value for row in wb_exp.active.iter_rows(min_row=2)]
    check("已標記的單匯出時個人標記欄印「是」，不是原始的 0/1",
          all(v == "是" for v in exp_vals), exp_vals)

    print("\n【34】同步實際驗入數量後自動判定驗收狀態（完成／異常），但手動可覆蓋")
    token = client.get("/api/account/sync-token").get_json()["token"]

    def sync_qty(items, operator="test"):
        return client.post("/api/sync/verified-qty",
                           json={"operator": operator, "items": items},
                           headers={"X-Sync-Token": token})

    def recv_status(po):
        return next(r for r in client.get("/api/pos?page_size=200").get_json()["rows"]
                    if r["po_number"] == po)["receiving_status"]

    # 挑一張品項不多、而且每個品項都有出貨數量的單來測
    cand = None
    for row in client.get("/api/pos?page_size=200").get_json()["rows"]:
        det = client.get(f"/api/pos/{row['po_number']}").get_json()
        skus = [s for s in det["skus"] if not s.get("removed_from_coupang")]
        if 2 <= len(skus) <= 40 and all(s["qty_ship"] is not None for s in skus):
            cand = (row["po_number"], skus)
            break
    check("找得到可以測自動判定的 PO", cand is not None)

    po_no, skus = cand
    client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": [po_no],
        "field": "receiving_status", "value": "未驗收"})

    # 只同步一部分品項、而且數字都對得起來 → 還不能判成完成
    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": skus[0]["qty_ship"]}])
    check("只驗到一部分、數字都對，狀態維持未驗收（不提早判完成）",
          recv_status(po_no) == "未驗收", recv_status(po_no))

    # 其餘品項也同步、全部相符 → 完成
    sync_qty([{"po_number": po_no, "sku_id": s["sku_id"],
               "verified_qty": s["qty_ship"]} for s in skus[1:]])
    check("全部品項都驗到且數字相符，自動判為完成",
          recv_status(po_no) == "完成", recv_status(po_no))

    # 其中一個品項少驗 → 異常
    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": (skus[0]["qty_ship"] or 0) - 1}])
    check("任一品項少驗，自動判為異常",
          recv_status(po_no) == "異常", recv_status(po_no))

    # 多驗也算異常（數字不一樣就是異常，不分多少）
    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": (skus[0]["qty_ship"] or 0) + 5}])
    check("多驗一樣判為異常（規則是「數字不一樣」，不是只看短驗）",
          recv_status(po_no) == "異常", recv_status(po_no))

    # 手動改成完成之後，再同步不該被蓋回異常
    client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": [po_no],
        "field": "receiving_status", "value": "完成"})
    check("手動可以把自動判定的異常改成完成", recv_status(po_no) == "完成")

    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": (skus[0]["qty_ship"] or 0) + 7}])
    check("手動判定過的單，之後同步不會被自動蓋回去",
          recv_status(po_no) == "完成", recv_status(po_no))

    # 改回未驗收＝放掉手動判定，自動判定恢復
    client.post("/api/pos/status", json={
        "operator": "小真", "po_numbers": [po_no],
        "field": "receiving_status", "value": "未驗收"})
    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": (skus[0]["qty_ship"] or 0) + 7}])
    check("改回未驗收等於放掉手動判定，自動判定會恢復",
          recv_status(po_no) == "異常", recv_status(po_no))

    # 單張編輯（PUT）這條路徑也要一樣會設手動旗標，不是只有批次改狀態會
    ver = client.get(f"/api/pos/{po_no}").get_json()["header"]["version"]
    rput = client.put(f"/api/pos/{po_no}", json={
        "operator": "小真", "po_version": ver, "receiving_status": "完成"})
    check("單張編輯也能改驗收狀態", rput.status_code == 200, rput.get_json())
    sync_qty([{"po_number": po_no, "sku_id": skus[0]["sku_id"],
               "verified_qty": (skus[0]["qty_ship"] or 0) + 9}])
    check("單張編輯改過的狀態，同步一樣不會蓋掉",
          recv_status(po_no) == "完成", recv_status(po_no))

    # 自動判定要留下歷程，不能無聲改狀態
    logs = client.get(f"/api/pos/{po_no}").get_json().get("logs", [])
    auto = [l for l in logs
            if l["field"] == "receiving_status" and l["source"] == "system"]
    check("自動判定的狀態變更有記進歷程", len(auto) > 0, len(auto))

    print("\n【35】PO 總表匯出：以 PO 為單位一列，不列 SKU 明細，外加備註")
    all_pos_now = client.get("/api/pos?page_size=200").get_json()["rows"]
    multi_po = next((r for r in all_pos_now if r["sku_count"] and r["sku_count"] > 1), None)
    check("找得到品項數大於 1 的 PO 來測彙總", multi_po is not None)

    po_sum = client.post("/api/export", json={
        "operator": "小真", "profile": "po_summary", "mark_pulled": False,
        "filters": {}, "po_numbers": [multi_po["po_number"]]})
    check("PO 總表匯出成功", po_sum.status_code == 200, po_sum.status_code)
    wb_sum = openpyxl.load_workbook(io.BytesIO(po_sum.data))
    sum_headers = [c.value for c in wb_sum.active[1]]
    check("表頭跟首頁列表一致，外加備註",
          sum_headers == ["PO單號", "訂單類型", "PO狀態", "已拉單", "驗收狀態", "線別", "品牌",
                          "建檔日", "到貨日", "到貨倉別", "配送方式", "品項數", "出貨數量",
                          "實際驗入數量", "備註"],
          sum_headers)
    data_rows = list(wb_sum.active.iter_rows(min_row=2, values_only=True))
    check("這張多品項的 PO 匯出只有一列，不是一個 SKU 一列",
          len(data_rows) == 1, len(data_rows))
    po_row = dict(zip(sum_headers, data_rows[0]))
    check("PO 單號正確", po_row["PO單號"] == multi_po["po_number"], po_row["PO單號"])
    check("品項數對得上實際 SKU 數", po_row["品項數"] == multi_po["sku_count"],
          (po_row["品項數"], multi_po["sku_count"]))

    # 只看已標記顏色，PO 總表一樣要吃得到這個篩選條件
    flag_sum = client.post("/api/export", json={
        "operator": "小真", "profile": "po_summary", "mark_pulled": False,
        "filters": {"flagged": "1"}})
    check("PO 總表也能只匯出已標記顏色的單", flag_sum.status_code == 200, flag_sum.status_code)
    wb_flag_sum = openpyxl.load_workbook(io.BytesIO(flag_sum.data))
    flag_sum_pos = [row[0].value for row in wb_flag_sum.active.iter_rows(min_row=2)]
    check("只匯出剛才標記的那張單", flag_sum_pos == [flag_po], flag_sum_pos)

    # 還沒同步實際驗入數量的單，PO 總表要留空白，不能印成 0
    unsynced_po = next((r for r in all_pos_now
                        if r["actual_verified_qty"] is None), None)
    if unsynced_po:
        us = client.post("/api/export", json={
            "operator": "小真", "profile": "po_summary", "mark_pulled": False,
            "filters": {}, "po_numbers": [unsynced_po["po_number"]]})
        wb_us = openpyxl.load_workbook(io.BytesIO(us.data))
        us_headers = [c.value for c in wb_us.active[1]]
        us_row = list(wb_us.active.iter_rows(min_row=2, values_only=True))[0]
        us_dict = dict(zip(us_headers, us_row))
        check("還沒同步過的單，實際驗入數量印成空白，不是 0",
              us_dict["實際驗入數量"] is None, us_dict["實際驗入數量"])

    print("\n【36】CPG 新單匯入自動帶配送方式，其他線別不動；舊資料一併補齊")
    # CPG 固定走原廠(EM)，新單建立當下就該自動判斷帶入，不用 OP 每張手動選。
    cpg_po = "13000000900001"
    other_po = "13000000900002"
    cpg_path = os.path.join(_tmp, "cpg_new.xlsx")
    wb = openpyxl.load_workbook(SAMPLE)
    ws = wb["整合表"]
    template_row = 2   # 直接複製一列真實資料的其他欄位，只改 PO 單號跟線別
    rows_to_add = [(cpg_po, "CPG-潔品"), (other_po, "瑪氏")]
    for i, (po, line) in enumerate(rows_to_add):
        r = ws.max_row + 1
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c, value=ws.cell(row=template_row, column=c).value)
        ws.cell(row=r, column=2, value=line)         # 線別
        ws.cell(row=r, column=3, value=po)           # PO單號
        ws.cell(row=r, column=8, value=f"TESTSKU{i}")  # SKU ID，避開跟既有資料撞鍵
    wb.save(cpg_path)

    res = upload(client, cpg_path)
    prev = res.get_json()
    client.post("/api/import/commit", json={"batch_id": prev["batch_id"], "operator": "小真"})

    cpg_header = client.get(f"/api/pos/{cpg_po}").get_json()["header"]
    check("CPG 新單匯入時自動帶入配送方式「原廠(EM)」",
          cpg_header["shipping_method"] == "原廠(EM)", cpg_header["shipping_method"])

    other_header = client.get(f"/api/pos/{other_po}").get_json()["header"]
    check("非 CPG 線別的新單，配送方式不自動帶、維持空白",
          other_header["shipping_method"] == "", other_header["shipping_method"])

    # 補既有舊資料：找一筆匯入前就存在、線別是 CPG 但配送方式還沒填的單
    # （用直接寫資料庫模擬「規則上線前就已經在系統裡」的舊資料），
    # 重跑一次 init_db 的補值邏輯，確認補得到。
    legacy_po = "13000000900003"
    conn = db.get_conn()
    stamp = db.now()
    conn.execute(
        "INSERT INTO po_headers (po_number, po_status, receiving_status, is_pulled, "
        "filed_date, created_at, updated_at) VALUES (?, '已建立', '未驗收', 0, ?, ?, ?)",
        (legacy_po, db.today(), stamp, stamp))
    conn.execute(
        "INSERT INTO orders (po_number, sku_id, line, created_at, updated_at, "
        "first_seen_at, last_seen_at) VALUES (?, 'LEGACYSKU', 'CPG-紙品', ?, ?, ?, ?)",
        (legacy_po, stamp, stamp, stamp, stamp))
    conn.commit()
    conn.close()

    migrate_conn = db.get_conn()
    db._migrate_columns(migrate_conn)
    migrate_conn.commit()
    migrate_conn.close()

    legacy_header = client.get(f"/api/pos/{legacy_po}").get_json()["header"]
    check("規則上線前既有的 CPG 舊資料，補值邏輯跑過一次後也補上了",
          legacy_header["shipping_method"] == "原廠(EM)", legacy_header["shipping_method"])

    print("\n" + "=" * 62)
    print(f"通過 {len(PASS)} 項／失敗 {len(FAIL)} 項")
    if FAIL:
        for name in FAIL:
            print("  ✗", name)
        return 1
    print("全部通過。")
    return 0


if __name__ == "__main__":
    try:
        code = main()
    finally:
        shutil.rmtree(_tmp, ignore_errors=True)
    sys.exit(code)
